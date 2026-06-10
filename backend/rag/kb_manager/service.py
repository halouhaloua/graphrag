import os
import glob
import shutil
import asyncio
import pathlib

from loguru import logger

from typing import List, Optional
from typing import List as _List

from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select as _select, delete as _delete

from rag.file_manager.ocr import _common_ocr_sync
from rag.kb_manager.schema import FileUploadResponse
from rag.kb_manager.model import KnowledgeBaseRole
from rag.kb_manager.model import KnowledgeBaseFile as KBFileModel

path = pathlib.Path(__file__).parent.parent
DATA_UPLOAD_DIR = path / "data" / "uploads"

DEFAULT_SCHEMA = {
    "Nodes": [
        "person", "location", "organization", "event",
        "object", "concept", "time_period", "creative_work",
    ],
    "Relations": [
        "is_a", "part_of", "located_in", "created_by", "used_by",
        "participates_in", "related_to", "belongs_to", "influences",
        "precedes", "arrives_in", "comparable_to",
    ],
    "Attributes": [
        "name", "date", "size", "type", "description",
        "status", "quantity", "value", "position", "duration", "time",
    ],
}


def _ensure_schema(schema_json: Optional[dict]) -> dict:
    return schema_json or DEFAULT_SCHEMA


def _clear_cache_files_sync(kb_id: str, file_id: str):
    try:
        cache_key = file_id
        upload_dir = DATA_UPLOAD_DIR / kb_id
        if upload_dir.exists():
            for fp in glob.glob(str(upload_dir / f"{cache_key}.*")):
                try:
                    if os.path.isfile(fp):
                        os.remove(fp)
                        logger.info(f"Removed uploaded file: {fp}")
                except Exception as e:
                    logger.warning(f"Failed to remove uploaded file {fp}: {e}")
            file_subdir = upload_dir / cache_key
            if file_subdir.is_dir():
                shutil.rmtree(file_subdir)
                logger.info(f"Removed upload subdirectory: {file_subdir}")
            try:
                if not any(upload_dir.iterdir()):
                    upload_dir.rmdir()
                    logger.info(f"Removed empty upload directory: {upload_dir}")
            except Exception:
                pass
        faiss_cache_dir = path / f"retriever/faiss_cache_new/{cache_key}"
        if os.path.exists(faiss_cache_dir):
            shutil.rmtree(faiss_cache_dir)
        chunk_file = path / f"output/chunks/{cache_key}.txt"
        if os.path.exists(chunk_file):
            os.remove(chunk_file)
        graph_file = path / f"output/graphs/{cache_key}_new.json"
        if os.path.exists(graph_file):
            os.remove(graph_file)
        for pattern in [
            path / f"output/logs/{cache_key}_*.log",
            path / f"output/chunks/{cache_key}_*",
            path / f"output/graphs/{cache_key}_*",
            path / f"output/chunks/{cache_key}",
            path / f"output/graphs/{cache_key}",
            path / f"output/logs/{cache_key}",
        ]:
            for fp in glob.glob(str(pattern)):
                try:
                    if os.path.isfile(fp):
                        os.remove(fp)
                    elif os.path.isdir(fp):
                        shutil.rmtree(fp)
                except Exception as e:
                    logger.warning(f"Failed to clear {fp}: {e}")
        logger.info(f"Cache cleanup completed for file: {file_id}")
    except Exception as e:
        logger.error(f"Error clearing cache files for {file_id}: {e}")


async def clear_cache_files(kb_id: str, file_id: str):
    loop = asyncio.get_event_loop()
    await loop.run_in_executor(None, _clear_cache_files_sync, kb_id, file_id)


def _extract_text_from_document_sync(file_path: str, file_ext: str) -> str:
    return _common_ocr_sync(file_path, file_ext)


async def extract_text_from_document(file_path: str, file_ext: str) -> str:
    loop = asyncio.get_event_loop()
    return await loop.run_in_executor(None, _extract_text_from_document_sync, file_path, file_ext)


def _extract_text_from_spreadsheet_sync(file_path: str, file_ext: str):
    try:
        if file_ext == ".csv":
            import csv
            text = ""
            with open(file_path, "r", encoding="utf-8", errors="ignore") as f:
                for row in csv.reader(f):
                    text += ", ".join(row) + "\n"
            return text
        elif file_ext in [".xls", ".xlsx"]:
            try:
                import pandas as pd
                return pd.read_excel(file_path).to_string(index=False)
            except ImportError:
                import openpyxl
                wb = openpyxl.load_workbook(file_path)
                text = ""
                for sheet in wb.sheetnames:
                    ws = wb[sheet]
                    text += f"Sheet: {sheet}\n"
                    for row in ws.iter_rows():
                        text += ", ".join(str(c.value) if c.value else "" for c in row) + "\n"
                    text += "\n"
                return text
    except Exception as e:
        return f"[Error: {str(e)}]"
    return ""


async def extract_text_from_spreadsheet(file_path: str, file_ext: str):
    loop = asyncio.get_event_loop()
    return await loop.run_in_executor(None, _extract_text_from_spreadsheet_sync, file_path, file_ext)


async def upload_files_to_kb(
    kb_id: str, files: List, schema_json: Optional[dict], db: AsyncSession
) -> List["FileUploadResponse"]:

    if not files:
        raise Exception("No files provided")
    effective_schema = _ensure_schema(schema_json)
    results = []
    for file in files:
        content_bytes = await file.read()
        file_ext = os.path.splitext(file.filename)[1].lower()
        file_size = len(content_bytes)
        from app.base_model import generate_nanoid
        file_id = generate_nanoid()
        upload_path = DATA_UPLOAD_DIR / kb_id
        upload_path.mkdir(parents=True, exist_ok=True)
        save_path = upload_path / f"{file_id}{file_ext}"
        save_path.write_bytes(content_bytes)
        if file_ext in [".csv", ".xls", ".xlsx"]:
            text = await extract_text_from_spreadsheet(str(save_path), file_ext)
        else:
            text = await extract_text_from_document(str(save_path), file_ext)
        text = text.replace('\x00', '')
        file_record = KBFileModel(
            id=file_id,
            kb_id=kb_id,
            filename=file.filename,
            content=text,
            file_type=file_ext if file_ext else ".txt",
            file_size=file_size,
            schema_json=effective_schema,
            has_graph=False,
        )
        db.add(file_record)
        await db.flush()
        await db.refresh(file_record)
        results.append(
            FileUploadResponse(
                id=file_record.id,
                filename=file.filename,
                file_type=file_ext if file_ext else ".txt",
                file_size=file_size,
                has_graph=False,
            )
        )
    await db.commit()
    return results


# ─── 角色/部门/用户权限 ───
class KnowledgeBasePermissionService:

    @classmethod
    def _get_role_ids(cls, user) -> _List[str]:
        """从 user 对象获取所有角色 ID，支持多角色"""
        from utils.context import get_current_user_info_from_context
        ctx = get_current_user_info_from_context()
        if ctx and ctx.get("role_ids"):
            return ctx["role_ids"]
        return [user.role_id] if user.role_id else []

    @classmethod
    def _get_dept_id(cls, user) -> Optional[str]:
        """从 user 对象获取部门 ID"""
        from utils.context import get_current_user_info_from_context
        ctx = get_current_user_info_from_context()
        if ctx and ctx.get("dept_id"):
            return ctx["dept_id"]
        return getattr(user, "dept_id", None)

    # ─── 角色维度 ───

    @classmethod
    async def get_role_kb_ids(cls, db: AsyncSession, role_id: str) -> _List[str]:
        query = _select(KnowledgeBaseRole.kb_id).where(
            KnowledgeBaseRole.role_id == role_id,
            KnowledgeBaseRole.is_deleted == False,
        )
        result = await db.execute(query)
        return [row[0] for row in result]

    @classmethod
    async def set_role_kbs(
        cls, db: AsyncSession, role_id: str, kb_ids: _List[str]
    ):
        await db.execute(
            _delete(KnowledgeBaseRole).where(
                KnowledgeBaseRole.role_id == role_id,
            )
        )
        for kb_id in kb_ids:
            db.add(KnowledgeBaseRole(role_id=role_id, kb_id=kb_id))
        await db.commit()

    # ─── 部门维度 ───

    @classmethod
    async def get_dept_kb_ids(cls, db: AsyncSession, dept_id: str) -> _List[str]:
        from rag.kb_manager.model import KnowledgeBaseDept
        query = _select(KnowledgeBaseDept.kb_id).where(
            KnowledgeBaseDept.dept_id == dept_id,
            KnowledgeBaseDept.is_deleted == False,
        )
        result = await db.execute(query)
        return [row[0] for row in result]

    @classmethod
    async def set_dept_kbs(
        cls, db: AsyncSession, dept_id: str, kb_ids: _List[str]
    ):
        from rag.kb_manager.model import KnowledgeBaseDept
        await db.execute(
            _delete(KnowledgeBaseDept).where(
                KnowledgeBaseDept.dept_id == dept_id,
            )
        )
        for kb_id in kb_ids:
            db.add(KnowledgeBaseDept(dept_id=dept_id, kb_id=kb_id))
        await db.commit()

    # ─── 用户维度 ───

    @classmethod
    async def get_user_kb_ids(cls, db: AsyncSession, user_id: str) -> _List[str]:
        from rag.kb_manager.model import KnowledgeBaseUser
        query = _select(KnowledgeBaseUser.kb_id).where(
            KnowledgeBaseUser.user_id == user_id,
            KnowledgeBaseUser.is_deleted == False,
        )
        result = await db.execute(query)
        return [row[0] for row in result]

    @classmethod
    async def set_user_kbs(
        cls, db: AsyncSession, user_id: str, kb_ids: _List[str]
    ):
        from rag.kb_manager.model import KnowledgeBaseUser
        await db.execute(
            _delete(KnowledgeBaseUser).where(
                KnowledgeBaseUser.user_id == user_id,
            )
        )
        for kb_id in kb_ids:
            db.add(KnowledgeBaseUser(user_id=user_id, kb_id=kb_id))
        await db.commit()

    # ─── 检查 / 查询（OR 逻辑：全员可见 OR 角色 OR 部门 OR 用户任一匹配） ───

    @classmethod
    async def _kb_is_public(cls, db: AsyncSession, kb_id: str) -> bool:
        """检查知识库是否设置为全员可见"""
        from rag.kb_manager.model import KnowledgeBase
        query = _select(KnowledgeBase.is_public).where(
            KnowledgeBase.id == kb_id,
            KnowledgeBase.is_deleted == False,
        )
        row = (await db.execute(query)).first()
        return row is not None and row[0] is True

    @classmethod
    async def check_kb_access(
        cls, db: AsyncSession, kb_id: str, user
    ) -> bool:
        if user.is_superuser:
            return True
        # 全员可见
        if await cls._kb_is_public(db, kb_id):
            return True
        # 角色维度
        role_ids = cls._get_role_ids(user)
        if role_ids:
            query = _select(KnowledgeBaseRole.id).where(
                KnowledgeBaseRole.role_id.in_(role_ids),
                KnowledgeBaseRole.kb_id == kb_id,
                KnowledgeBaseRole.is_deleted == False,
            )
            if (await db.execute(query)).first() is not None:
                return True
        # 部门维度
        dept_id = cls._get_dept_id(user)
        if dept_id:
            from rag.kb_manager.model import KnowledgeBaseDept
            query = _select(KnowledgeBaseDept.id).where(
                KnowledgeBaseDept.dept_id == dept_id,
                KnowledgeBaseDept.kb_id == kb_id,
                KnowledgeBaseDept.is_deleted == False,
            )
            if (await db.execute(query)).first() is not None:
                return True
        # 用户维度
        from rag.kb_manager.model import KnowledgeBaseUser
        query = _select(KnowledgeBaseUser.id).where(
            KnowledgeBaseUser.user_id == user.id,
            KnowledgeBaseUser.kb_id == kb_id,
            KnowledgeBaseUser.is_deleted == False,
        )
        return (await db.execute(query)).first() is not None

    @classmethod
    async def get_user_visible_kb_ids(
        cls, db: AsyncSession, user
    ) -> Optional[_List[str]]:
        if user.is_superuser:
            return None
        visible: set[str] = set()
        # 全员可见
        from rag.kb_manager.model import KnowledgeBase
        query = _select(KnowledgeBase.id).where(
            KnowledgeBase.is_public == True,
            KnowledgeBase.is_deleted == False,
        )
        for row in (await db.execute(query)):
            visible.add(row[0])
        # 角色维度
        role_ids = cls._get_role_ids(user)
        if role_ids:
            query = _select(KnowledgeBaseRole.kb_id).where(
                KnowledgeBaseRole.role_id.in_(role_ids),
                KnowledgeBaseRole.is_deleted == False,
            )
            for row in (await db.execute(query)):
                visible.add(row[0])
        # 部门维度
        dept_id = cls._get_dept_id(user)
        if dept_id:
            from rag.kb_manager.model import KnowledgeBaseDept
            query = _select(KnowledgeBaseDept.kb_id).where(
                KnowledgeBaseDept.dept_id == dept_id,
                KnowledgeBaseDept.is_deleted == False,
            )
            for row in (await db.execute(query)):
                visible.add(row[0])
        # 用户维度
        from rag.kb_manager.model import KnowledgeBaseUser
        query = _select(KnowledgeBaseUser.kb_id).where(
            KnowledgeBaseUser.user_id == user.id,
            KnowledgeBaseUser.is_deleted == False,
        )
        for row in (await db.execute(query)):
            visible.add(row[0])
        return list(visible)

    # ─── 批量设置 KB 的所有权限 ───

    @classmethod
    async def set_kb_permissions(
        cls, db: AsyncSession, kb_id: str,
        role_ids: Optional[_List[str]] = None,
        dept_ids: Optional[_List[str]] = None,
        user_ids: Optional[_List[str]] = None,
        auto_commit: bool = True,
    ):
        from rag.kb_manager.model import KnowledgeBaseDept, KnowledgeBaseUser
        if role_ids is not None:
            await db.execute(
                _delete(KnowledgeBaseRole).where(
                    KnowledgeBaseRole.kb_id == kb_id,
                )
            )
            for rid in role_ids:
                db.add(KnowledgeBaseRole(role_id=rid, kb_id=kb_id))
        if dept_ids is not None:
            await db.execute(
                _delete(KnowledgeBaseDept).where(
                    KnowledgeBaseDept.kb_id == kb_id,
                )
            )
            for did in dept_ids:
                db.add(KnowledgeBaseDept(dept_id=did, kb_id=kb_id))
        if user_ids is not None:
            await db.execute(
                _delete(KnowledgeBaseUser).where(
                    KnowledgeBaseUser.kb_id == kb_id,
                )
            )
            for uid in user_ids:
                db.add(KnowledgeBaseUser(user_id=uid, kb_id=kb_id))
        if auto_commit:
            await db.commit()

    @classmethod
    async def get_kb_permissions(
        cls, db: AsyncSession, kb_id: str
    ) -> dict:
        from rag.kb_manager.model import KnowledgeBaseDept, KnowledgeBaseUser
        # 全员可见
        is_public = await cls._kb_is_public(db, kb_id)
        # 角色
        result = await db.execute(
            _select(KnowledgeBaseRole.role_id).where(
                KnowledgeBaseRole.kb_id == kb_id,
                KnowledgeBaseRole.is_deleted == False,
            )
        )
        role_ids = [row[0] for row in result]
        # 部门
        result = await db.execute(
            _select(KnowledgeBaseDept.dept_id).where(
                KnowledgeBaseDept.kb_id == kb_id,
                KnowledgeBaseDept.is_deleted == False,
            )
        )
        dept_ids = [row[0] for row in result]
        # 用户
        result = await db.execute(
            _select(KnowledgeBaseUser.user_id).where(
                KnowledgeBaseUser.kb_id == kb_id,
                KnowledgeBaseUser.is_deleted == False,
            )
        )
        user_ids = [row[0] for row in result]
        return {
            "role_ids": role_ids,
            "dept_ids": dept_ids,
            "user_ids": user_ids,
            "is_public": is_public,
        }
