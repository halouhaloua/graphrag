#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
表单数据操作服务（异步版本）
支持动态操作不同数据表，适配 PostgreSQL、MySQL
"""
import asyncio
import json
import logging
import uuid
from datetime import date, datetime
from decimal import Decimal
from io import BytesIO
from typing import Any, Dict, List, Optional, Set, Tuple

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from sqlalchemy import select, text
from sqlalchemy.ext.asyncio import AsyncSession

from online_dev.form_manager.model import FormMeta, FormSubTable
from online_dev.form_data_manager.dynamic_sql_builder import DynamicSQLBuilder
from core.database_manager.service import AsyncDatabaseManagerService
from app.timezone import APP_TIMEZONE

logger = logging.getLogger(__name__)


def get_max_import_rows() -> int:
    """根据服务器物理内存计算最大导入/导出行数"""
    import psutil
    total_gb = psutil.virtual_memory().total / (1024 ** 3)
    thresholds = [(2, 100_000), (4, 200_000), (8, 600_000), (16, 1_048_575)]
    if total_gb <= 2:
        return 100_000
    for i in range(len(thresholds) - 1):
        low_gb, low_rows = thresholds[i]
        high_gb, high_rows = thresholds[i + 1]
        if total_gb <= high_gb:
            ratio = (total_gb - low_gb) / (high_gb - low_gb)
            return int(low_rows + ratio * (high_rows - low_rows))
    return 1_048_575


def get_server_memory_gb() -> float:
    """获取服务器物理内存（GB）"""
    import psutil
    return round(psutil.virtual_memory().total / (1024 ** 3), 1)


MAX_IMPORT_EXPORT_ROWS = get_max_import_rows()
SERVER_MEMORY_GB = get_server_memory_gb()
logger.info(f"服务器内存: {SERVER_MEMORY_GB}GB, 导入导出行数上限: {MAX_IMPORT_EXPORT_ROWS}")


class FormDataException(Exception):
    """表单数据操作异常"""
    pass


class FormDataService:
    """
    表单数据操作服务（异步版本）
    
    根据表单配置动态操作数据表，支持主表和子表的 CRUD 操作
    """

    def __init__(self, form_meta: FormMeta, sub_tables: List[FormSubTable], db_type: str = "postgresql"):
        """
        初始化服务
        
        Args:
            form_meta: 表单元数据
            sub_tables: 子表配置列表
            db_type: 数据库类型
        """
        self.form_meta = form_meta
        self.sub_tables = sub_tables
        self.db_type = db_type
        self.sql_builder = DynamicSQLBuilder(db_type)
        # 缓存 JSON 配置，避免 session 过期后懒加载导致 MissingGreenlet
        self._form_config = form_meta.form_config or {}
        self._list_config = form_meta.list_config or {}
    
    async def _resolve_schema(self, db: AsyncSession, schema: str) -> Optional[str]:
        """
        解析 schema 中的变量
        
        如果 schema 包含 {{application_code}} 变量，则从表单关联的应用中获取 application_code
        
        Args:
            db: 数据库会话
            schema: 原始 schema 配置（可能包含变量）
        
        Returns:
            解析后的 schema 值
        """
        if not schema:
            return None
        
        # 检查是否包含变量语法
        if '{{' in schema and '}}' in schema:
            # 提取变量名
            import re
            match = re.search(r'\{\{([^}]+)\}\}', schema)
            if match:
                var_name = match.group(1).strip()
                
                # 如果是 application_code 变量，从应用中获取
                if 'application_code' in var_name and self.form_meta.application_id:
                    from core.application.model import Application
                    result = await db.execute(
                        select(Application.code).where(Application.id == self.form_meta.application_id)
                    )
                    app_code = result.scalar_one_or_none()
                    if app_code:
                        # 替换变量
                        return schema.replace(f'{{{{{var_name}}}}}', app_code)
        
        return schema

    @classmethod
    async def create_service(cls, db: AsyncSession, form_code: str) -> "FormDataService":
        """
        工厂方法：创建服务实例
        
        Args:
            db: 数据库会话
            form_code: 表单编码
        """
        # 加载表单元数据
        stmt = select(FormMeta).where(
            FormMeta.code == form_code,
            FormMeta.is_deleted == False
        )
        result = await db.execute(stmt)
        form_meta = result.scalar_one_or_none()

        if not form_meta:
            raise FormDataException(f"表单不存在: {form_code}")

        # 加载子表配置
        sub_stmt = select(FormSubTable).where(
            FormSubTable.form_id == form_meta.id,
            FormSubTable.is_deleted == False
        ).order_by(FormSubTable.sort)
        sub_result = await db.execute(sub_stmt)
        sub_tables = list(sub_result.scalars().all())

        # 获取数据库类型
        from core.database_manager.service import parse_database_url
        from app.config import settings
        db_info = parse_database_url(settings.DATABASE_URL)
        db_type = db_info['db_type'] if db_info else "postgresql"

        return cls(form_meta, sub_tables, db_type)

    # ============ 字段白名单 ============

    def _get_allowed_fields(self, table_type: str = "main", table_name: str = None) -> Set[str]:
        """
        从表单配置中提取允许的字段（白名单）
        """
        fields = set()
        form_config = self._form_config
        items = form_config.get("items", [])

        def traverse(item_list: List[Dict], in_sub_table: str = None):
            """递归遍历表单项"""
            for item in item_list:
                item_type = item.get("type", "")
                field = item.get("field", "")

                # 子表单组件
                if item_type == "sub-table":
                    sub_table_name = field
                    children = item.get("children", [])
                    traverse(children, sub_table_name)
                    continue

                # 布局/展示组件，递归处理或跳过
                if item_type in ("grid", "tabs", "collapse", "steps", "divider", "alert", "timeline", "text", "html", "spacer", "title"):
                    if item.get("columns"):
                        for col in item["columns"]:
                            traverse(col.get("children", []), in_sub_table)
                    if item.get("items"):
                        for sub_item in item["items"]:
                            traverse(sub_item.get("children", []), in_sub_table)
                    continue

                # 普通字段（排除虚拟字段，虚拟字段不对应数据库列）
                if field:
                    props = item.get("props", {})
                    if props.get("isVirtualField"):
                        continue
                    if table_type == "main" and in_sub_table is None:
                        fields.add(field)
                    elif table_type == "sub" and in_sub_table == table_name:
                        fields.add(field)

        traverse(items)

        # 始终允许 id 字段
        fields.add("id")

        return fields

    def _filter_fields(self, data: Dict[str, Any], allowed_fields: Set[str]) -> Dict[str, Any]:
        """过滤字段，只保留白名单中的字段"""
        return {k: v for k, v in data.items() if k in allowed_fields}

    def _fill_system_fields_for_create(self, data: Dict[str, Any]) -> Dict[str, Any]:
        """
        填充新增时的系统字段
        sys_create_datetime, sys_update_datetime, sys_creator_id, sys_modifier_id, sys_dept_id, sort
        """
        from utils.context import get_current_user_info_from_context
        
        now = datetime.now()
        user_info = get_current_user_info_from_context()
        
        user_id = user_info.get('user_id') if user_info else None
        dept_id = user_info.get('dept_id') if user_info else None
        
        # 设置创建时间和更新时间
        data['sys_create_datetime'] = now
        data['sys_update_datetime'] = now
        
        # 设置创建人和修改人
        if user_id:
            data['sys_creator_id'] = user_id
            data['sys_modifier_id'] = user_id
        
        # 设置部门
        if dept_id:
            data['sys_dept_id'] = dept_id
        
        # 设置排序字段默认值（如果表中有 sort 字段且未提供值）
        if 'sort' not in data or data.get('sort') is None:
            data['sort'] = 0
        
        return data

    def _fill_system_fields_for_update(self, data: Dict[str, Any]) -> Dict[str, Any]:
        """
        填充更新时的系统字段
        sys_update_datetime, sys_modifier_id
        """
        from utils.context import get_current_user_info_from_context
        
        now = datetime.now()
        user_info = get_current_user_info_from_context()
        
        user_id = user_info.get('user_id') if user_info else None
        
        # 设置更新时间
        data['sys_update_datetime'] = now
        
        # 设置修改人
        if user_id:
            data['sys_modifier_id'] = user_id
        
        # 移除不应该被更新的系统字段
        data.pop('sys_create_datetime', None)
        data.pop('sys_creator_id', None)
        data.pop('sys_dept_id', None)
        
        return data

    def _apply_data_scope_filter(
            self,
            filters: Dict[str, Any],
            data_scope: Dict[str, Any]
    ) -> Dict[str, Any]:
        """
        应用数据权限过滤条件
        
        Args:
            filters: 现有过滤条件
            data_scope: 数据权限配置，格式：
                {
                    'filter_type': 'all' | 'self' | 'dept' | 'dept_and_children' | 'custom',
                    'scope': 0-4,
                    'user_id': str | None,
                    'dept_id': str | None,
                    'dept_ids': List[str] | None
                }
        
        Returns:
            合并后的过滤条件
        """
        if not filters:
            filters = {}
        
        filter_type = data_scope.get('filter_type', 'all')
        
        if filter_type == 'all':
            # 全部数据，不添加过滤条件
            return filters
        elif filter_type == 'self':
            # 仅本人数据：需要表中有 sys_creator_id 字段
            user_id = data_scope.get('user_id')
            if user_id:
                filters['sys_creator_id'] = user_id
        elif filter_type == 'dept':
            # 本部门数据：需要表中有 dept_id 字段
            dept_id = data_scope.get('dept_id')
            if dept_id:
                filters['dept_id'] = dept_id
        elif filter_type == 'dept_and_children':
            # 本部门及下级：需要表中有 dept_id 字段
            dept_ids = data_scope.get('dept_ids', [])
            if dept_ids:
                filters['dept_id'] = {"type": "in", "value": dept_ids}
        elif filter_type == 'custom':
            # 自定义部门
            dept_ids = data_scope.get('dept_ids', [])
            if dept_ids:
                filters['dept_id'] = {"type": "in", "value": dept_ids}
        
        return filters

    async def _apply_data_scope_filter_safe(
            self,
            db: AsyncSession,
            table: str,
            schema: Optional[str],
            database: Optional[str],
            filters: Dict[str, Any],
            data_scope: Dict[str, Any]
    ) -> Dict[str, Any]:
        """
        安全地应用数据权限过滤（检查字段是否存在）
        
        如果表中不存在所需的字段，则跳过该过滤条件
        """
        if not filters:
            filters = {}
        
        filter_type = data_scope.get('filter_type', 'all')
        
        if filter_type == 'all':
            # 全部数据，不添加过滤条件
            return filters
        
        # 获取表的所有列
        columns = await self._get_table_columns(db, table, schema, database)
        
        if filter_type == 'self':
            # 仅本人数据：需要表中有 sys_creator_id 字段
            # 注意：字段为空时所有人可见（使用 eq_or_null）
            if 'sys_creator_id' in columns:
                user_id = data_scope.get('user_id')
                if user_id:
                    filters['sys_creator_id'] = {"type": "eq_or_null", "value": user_id}
            else:
                logger.warning(f"表 {table} 中不存在 sys_creator_id 字段，跳过数据权限过滤")
        elif filter_type == 'dept':
            # 本部门数据：需要表中有 dept_id 或 sys_dept_id 字段
            # 注意：字段为空时所有人可见（使用 eq_or_null）
            dept_field = 'sys_dept_id' if 'sys_dept_id' in columns else ('dept_id' if 'dept_id' in columns else None)
            if dept_field:
                dept_id = data_scope.get('dept_id')
                if dept_id:
                    filters[dept_field] = {"type": "eq_or_null", "value": dept_id}
            else:
                logger.warning(f"表 {table} 中不存在 dept_id 或 sys_dept_id 字段，跳过数据权限过滤")
        elif filter_type == 'dept_and_children':
            # 本部门及下级：需要表中有 dept_id 或 sys_dept_id 字段
            # 注意：字段为空时所有人可见（使用 in_or_null）
            dept_field = 'sys_dept_id' if 'sys_dept_id' in columns else ('dept_id' if 'dept_id' in columns else None)
            if dept_field:
                dept_ids = data_scope.get('dept_ids', [])
                if dept_ids:
                    filters[dept_field] = {"type": "in_or_null", "value": dept_ids}
                else:
                    # 如果没有部门ID，只返回字段为空的记录
                    filters[dept_field] = {"type": "null"}
            else:
                logger.warning(f"表 {table} 中不存在 dept_id 或 sys_dept_id 字段，跳过数据权限过滤")
        elif filter_type == 'custom':
            # 自定义部门
            # 注意：字段为空时所有人可见（使用 in_or_null）
            dept_field = 'sys_dept_id' if 'sys_dept_id' in columns else ('dept_id' if 'dept_id' in columns else None)
            if dept_field:
                dept_ids = data_scope.get('dept_ids', [])
                if dept_ids:
                    filters[dept_field] = {"type": "in_or_null", "value": dept_ids}
                else:
                    # 如果没有部门ID，只返回字段为空的记录
                    filters[dept_field] = {"type": "null"}
            else:
                logger.warning(f"表 {table} 中不存在 dept_id 或 sys_dept_id 字段，跳过数据权限过滤")
        
        return filters

    # ============ 数据库操作 ============

    async def _execute_query(self, db: AsyncSession, sql: str, params: Any = None) -> List[Dict[str, Any]]:
        """执行查询"""
        # 处理参数：字典直接使用，None 转为空字典
        if params is None:
            params = {}
        result = await db.execute(text(sql), params)
        rows = result.fetchall()
        columns = result.keys()
        return [dict(zip(columns, row)) for row in rows]

    async def _execute_command(self, db: AsyncSession, sql: str, params: Any = None) -> int:
        """执行命令，返回影响行数"""
        # 处理参数：字典直接使用，None 转为空字典
        if params is None:
            params = {}
        result = await db.execute(text(sql), params)
        return result.rowcount

    async def _get_table_columns(
            self,
            db: AsyncSession,
            table: str,
            schema: Optional[str],
            database: Optional[str]
    ) -> List[str]:
        """获取表的所有列名"""
        try:
            if self.sql_builder.db_type == "postgresql":
                schema_name = schema or "public"
                sql = """
                    SELECT column_name 
                    FROM information_schema.columns 
                    WHERE table_schema = :schema 
                    AND table_name = :table
                """
                params = {"schema": schema_name, "table": table}
            elif self.sql_builder.db_type == "mysql":
                db_name = database or "information_schema"
                sql = """
                    SELECT column_name 
                    FROM information_schema.columns 
                    WHERE table_schema = :database 
                    AND table_name = :table
                """
                params = {"database": db_name, "table": table}
            else:
                return []
            
            result = await self._execute_query(db, sql, params)
            return [row["column_name"] for row in result]
        except Exception as e:
            logger.error(f"获取表列信息失败: {str(e)}")
            return []

    async def _check_column_exists(
            self,
            db: AsyncSession,
            table: str,
            schema: Optional[str],
            database: Optional[str],
            column_name: str
    ) -> bool:
        """检查表中是否存在指定列"""
        columns = await self._get_table_columns(db, table, schema, database)
        return column_name in columns

    async def _validate_sort_field(
            self,
            db: AsyncSession,
            table: str,
            schema: Optional[str],
            database: Optional[str],
            sort_field: str
    ) -> Optional[str]:
        """验证排序字段是否存在于表中，如果不存在则返回 None"""
        try:
            # 构建表名
            table_name = self.sql_builder.build_table_name(table, schema, database)
            
            # 获取表的列信息
            columns = await self._get_table_columns(db, table, schema, database)
            if not columns:
                # 无法获取列信息，直接返回字段（不验证）
                return sort_field
            
            # 检查排序字段是否存在
            if sort_field in columns:
                return sort_field
            
            logger.warning(f"排序字段 {sort_field} 不存在于表 {table_name} 中，可用字段: {columns}")
            return None
            
        except Exception as e:
            logger.error(f"验证排序字段失败: {str(e)}")
            # 验证失败时返回 None，使用默认排序
            return None

    # ============ 查询操作 ============

    async def list(
            self,
            db: AsyncSession,
            page: int = 1,
            page_size: int = 20,
            filters: Dict[str, Any] = None,
            sort_list: List[Dict[str, str]] = None,
            data_scope: Dict[str, Any] = None,
            search: str = None,
            search_fields: List[str] = None
    ) -> Dict[str, Any]:
        """分页查询列表（仅主表，支持数据权限过滤）
        
        Args:
            sort_list: 排序列表，格式为 [{"field": "name", "order": "asc"}, ...]
            data_scope: 数据权限过滤配置
            search: 搜索关键词
            search_fields: 搜索字段列表（多字段模糊搜索，OR 关系）
        """
        import time
        _t0 = time.perf_counter()

        table = self.form_meta.main_table
        schema = await self._resolve_schema(db, self.form_meta.main_table_schema) or None
        database = self.form_meta.main_table_database or None

        # 应用数据权限过滤（先检查字段是否存在）
        if data_scope:
            filters = await self._apply_data_scope_filter_safe(db, table, schema, database, filters, data_scope)

        _t1 = time.perf_counter()
        logger.info(f"[list 耗时] 数据权限过滤: {(_t1 - _t0) * 1000:.1f}ms")

        # 处理关联字段的名称搜索（转换为 ID 搜索）
        if filters:
            filters = await self._convert_relation_name_to_id(db, filters)
            # 转换日期字符串为 date 对象（PostgreSQL asyncpg 需要）
            filters = self._convert_date_strings(filters)

        _t2 = time.perf_counter()
        logger.info(f"[list 耗时] 关联字段转换+日期转换: {(_t2 - _t1) * 1000:.1f}ms")

        # 处理多字段搜索（OR 关系）
        if search and search_fields:
            if not filters:
                filters = {}
            # 使用特殊的 __search__ 键来标记多字段搜索
            filters['__search__'] = {
                'keyword': search,
                'fields': search_fields
            }

        # 构建 ORDER BY（支持多字段排序）
        order_by = None
        if sort_list:
            order_clauses = []
            for sort_item in sort_list:
                sort_field = sort_item.get('field')
                sort_order = sort_item.get('order', 'desc')
                
                if not sort_field:
                    continue
                
                # 验证排序字段是否存在于表中
                valid_sort_field = await self._validate_sort_field(db, table, schema, database, sort_field)
                if valid_sort_field:
                    direction = "DESC" if sort_order.lower() == "desc" else "ASC"
                    order_clauses.append(f"{self.sql_builder.quote_identifier(valid_sort_field)} {direction}")
                else:
                    logger.warning(f"排序字段 {sort_field} 不存在于表 {table} 中，跳过")
            
            if order_clauses:
                order_by = ", ".join(order_clauses)

        _t3 = time.perf_counter()
        logger.info(f"[list 耗时] 排序字段验证: {(_t3 - _t2) * 1000:.1f}ms")

        # 查询总数
        count_sql, count_params = self.sql_builder.build_count(
            table, schema, database, filters
        )
        count_result = await self._execute_query(db, count_sql, count_params)
        total = count_result[0]["total"] if count_result else 0

        _t4 = time.perf_counter()
        logger.info(f"[list 耗时] COUNT 查询 (total={total}): {(_t4 - _t3) * 1000:.1f}ms | SQL: {count_sql}")

        # 查询数据
        offset = (page - 1) * page_size
        data_sql, data_params = self.sql_builder.build_select(
            table=table,
            schema=schema,
            database=database,
            where=filters,
            order_by=order_by,
            limit=page_size,
            offset=offset
        )

        rows = await self._execute_query(db, data_sql, data_params)

        _t5 = time.perf_counter()
        logger.info(f"[list 耗时] 数据查询 (rows={len(rows)}): {(_t5 - _t4) * 1000:.1f}ms | SQL: {data_sql}")

        # 处理特殊类型
        items = [self._serialize_row(row) for row in rows]

        _t6 = time.perf_counter()
        logger.info(f"[list 耗时] 序列化: {(_t6 - _t5) * 1000:.1f}ms")

        # 填充关联字段的显示名称
        relation_fields = self._get_relation_fields()
        logger.debug(f"识别到的关联字段: {relation_fields}")
        if relation_fields and items:
            items = await self._fill_relation_display_names(db, items, relation_fields)

        _t7 = time.perf_counter()
        logger.info(f"[list 耗时] 填充关联字段显示名称 (fields={len(relation_fields)}): {(_t7 - _t6) * 1000:.1f}ms")

        # 填充虚拟字段的值（基于值关联配置）
        items = await self._fill_virtual_fields(db, items, relation_fields, context="list")

        _t8 = time.perf_counter()
        logger.info(f"[list 耗时] 填充虚拟字段: {(_t8 - _t7) * 1000:.1f}ms")

        # 应用字段权限过滤
        items = await self.apply_field_permissions(items, db)

        _t9 = time.perf_counter()
        logger.info(f"[list 耗时] 字段权限过滤: {(_t9 - _t8) * 1000:.1f}ms")
        logger.info(f"[list 耗时] ===== 总耗时: {(_t9 - _t0) * 1000:.1f}ms (table={table}, page={page}, page_size={page_size}, total={total}) =====")

        return {
            "items": items,
            "total": total,
            "page": page,
            "page_size": page_size
        }

    async def get_tree_children(
            self,
            db: AsyncSession,
            parent_id: str = None,
            parent_field: str = "parent_id",
            data_scope: Dict[str, Any] = None
    ) -> List[Dict[str, Any]]:
        """获取树形数据的子节点（用于懒加载模式）
        
        Args:
            parent_id: 父节点ID，为空时获取根节点
            parent_field: 父节点字段名，默认为 parent_id
            data_scope: 数据权限过滤配置
        
        Returns:
            子节点列表，每个节点包含 has_children 标记
        """
        table = self.form_meta.main_table
        schema = await self._resolve_schema(db, self.form_meta.main_table_schema) or None
        database = self.form_meta.main_table_database or None

        # 验证 parent_field 是否存在
        valid_parent_field = await self._validate_sort_field(db, table, schema, database, parent_field)
        if not valid_parent_field:
            raise FormDataException(f"父节点字段不存在: {parent_field}")

        # 构建过滤条件
        filters = {}
        if parent_id:
            # 获取指定父节点的子节点
            filters[parent_field] = {"type": "eq", "value": parent_id}
        else:
            # 获取根节点（parent_field 为空或 NULL）
            filters[parent_field] = {"type": "null"}

        # 应用数据权限过滤
        if data_scope:
            filters = await self._apply_data_scope_filter_safe(db, table, schema, database, filters, data_scope)

        # 查询数据（按 sort 字段排序）
        order_by = f"{self.sql_builder.quote_identifier('sort')} ASC"
        
        data_sql, data_params = self.sql_builder.build_select(
            table=table,
            schema=schema,
            database=database,
            where=filters,
            order_by=order_by
        )

        rows = await self._execute_query(db, data_sql, data_params)
        items = [self._serialize_row(row) for row in rows]

        # 填充关联字段的显示名称
        relation_fields = self._get_relation_fields()
        if relation_fields and items:
            items = await self._fill_relation_display_names(db, items, relation_fields)

        # 为每个节点检查是否有子节点
        if items:
            table_name = self.sql_builder.build_table_name(table, schema, database)
            quoted_parent_field = self.sql_builder.quote_identifier(parent_field)
            
            # 检查表是否有 is_deleted 字段
            has_is_deleted = await self._check_column_exists(db, table, schema, database, "is_deleted")
            
            # 批量查询所有节点的子节点数量
            ids = [item.get('id') for item in items if item.get('id')]
            if ids:
                # 构建 IN 查询
                placeholders = ", ".join([f":id_{i}" for i in range(len(ids))])
                where_clause = f"{quoted_parent_field} IN ({placeholders})"
                if has_is_deleted:
                    where_clause += " AND is_deleted = false"
                
                count_sql = f"""
                    SELECT {quoted_parent_field} as parent_id, COUNT(*) as child_count
                    FROM {table_name}
                    WHERE {where_clause}
                    GROUP BY {quoted_parent_field}
                """
                count_params = {f"id_{i}": id_val for i, id_val in enumerate(ids)}
                
                try:
                    count_result = await self._execute_query(db, count_sql, count_params)
                    # 构建父ID到子节点数量的映射
                    child_count_map = {row['parent_id']: row['child_count'] for row in count_result}
                    
                    # 为每个节点设置 has_children 标记
                    for item in items:
                        item_id = item.get('id')
                        item['has_children'] = child_count_map.get(item_id, 0) > 0
                except Exception as e:
                    logger.warning(f"查询子节点数量失败: {e}")
                    # 如果查询失败，默认设置为 False
                    for item in items:
                        item['has_children'] = False

        return items

    async def get_field_values(
            self,
            db: AsyncSession,
            field_name: str,
            page: int = 1,
            page_size: int = 20,
            search: str = None
    ) -> Dict[str, Any]:
        """
        获取指定字段的唯一值列表（用于过滤选项）
        
        Args:
            db: 数据库会话
            field_name: 字段名
            page: 页码
            page_size: 每页数量
            search: 搜索关键词（模糊匹配）
        
        Returns:
            {
                "items": [{"value": "值", "label": "显示文本", "count": 数量}],
                "total": 总数,
                "hasMore": 是否有更多
            }
        """
        table = self.form_meta.main_table
        schema = await self._resolve_schema(db, self.form_meta.main_table_schema) or None
        database = self.form_meta.main_table_database or None
        
        # 验证字段是否存在
        valid_field = await self._validate_sort_field(db, table, schema, database, field_name)
        if not valid_field:
            raise FormDataException(f"字段不存在: {field_name}")
        
        # 构建表名
        table_name = self.sql_builder.build_table_name(table, schema, database)
        quoted_field = self.sql_builder.quote_identifier(field_name)
        
        # 构建 WHERE 条件（只检查 NOT NULL，不检查空字符串，因为非字符串类型会报错）
        where_clause = f"{quoted_field} IS NOT NULL"
        params = {}
        
        if search:
            if self.sql_builder.db_type == "postgresql":
                where_clause += f" AND CAST({quoted_field} AS TEXT) ILIKE :search"
            else:
                where_clause += f" AND CAST({quoted_field} AS CHAR) LIKE :search"
            params["search"] = f"%{search}%"
        
        # 检查表是否有 is_deleted 字段
        has_is_deleted = await self._check_column_exists(db, table, schema, database, "is_deleted")
        if has_is_deleted:
            where_clause += " AND is_deleted = false"
        
        # 查询总数
        count_sql = f"""
            SELECT COUNT(DISTINCT {quoted_field}) as total
            FROM {table_name}
            WHERE {where_clause}
        """
        count_result = await self._execute_query(db, count_sql, params)
        total = count_result[0]["total"] if count_result else 0
        
        # 查询唯一值（带分页）
        offset = (page - 1) * page_size
        data_sql = f"""
            SELECT {quoted_field} as value, COUNT(*) as count
            FROM {table_name}
            WHERE {where_clause}
            GROUP BY {quoted_field}
            ORDER BY count DESC, {quoted_field} ASC
            LIMIT :limit OFFSET :offset
        """
        params["limit"] = page_size
        params["offset"] = offset
        
        rows = await self._execute_query(db, data_sql, params)
        
        # 处理结果，获取显示标签
        items = []
        field_options = self._get_field_options(field_name)
        
        for row in rows:
            value = row["value"]
            # 序列化特殊类型
            if isinstance(value, (datetime, date)):
                value = value.isoformat()
            elif isinstance(value, Decimal):
                value = float(value)
            elif isinstance(value, uuid.UUID):
                value = str(value)
            
            # 获取显示标签
            label = str(value)
            if field_options:
                for opt in field_options:
                    if str(opt.get("value")) == str(value):
                        label = opt.get("label", str(value))
                        break
            
            items.append({
                "value": value,
                "label": label,
                "count": row["count"]
            })
        
        return {
            "items": items,
            "total": total,
            "hasMore": (page * page_size) < total
        }

    def _get_unique_check_fields(self) -> List[str]:
        """
        从表单配置中提取启用了唯一性校验的主表字段名列表
        
        支持两种配置方式：
        1. 表单设计器中的 props.uniqueCheck（字段级别）
        2. 数据库设计中的 tableConfigs[].fields[].uniqueCheck（表级别）
        """
        unique_fields = set()
        form_config = self._form_config
        
        # 方式1：从表单设计器的 items 中获取
        items = form_config.get("items", [])

        def traverse(item_list: List[Dict], in_sub_table: str = None):
            for item in item_list:
                item_type = item.get("type", "")
                field = item.get("field", "")

                if item_type == "sub-table":
                    traverse(item.get("children", []), field)
                    continue

                if item_type in ("grid", "tabs", "collapse", "steps", "divider", "alert", "timeline", "text", "html", "spacer", "title"):
                    if item.get("columns"):
                        for col in item["columns"]:
                            traverse(col.get("children", []), in_sub_table)
                    if item.get("items"):
                        for sub_item in item["items"]:
                            traverse(sub_item.get("children", []), in_sub_table)
                    continue

                if field and in_sub_table is None:
                    props = item.get("props", {})
                    if props.get("uniqueCheck"):
                        unique_fields.add(field)

        traverse(items)
        
        # 方式2：从数据库设计的 tableConfigs 中获取（主表字段）
        table_configs = form_config.get("tableConfigs", [])
        for table_config in table_configs:
            if table_config.get("type") == "main":
                for field in table_config.get("fields", []):
                    if field.get("uniqueCheck"):
                        unique_fields.add(field.get("name"))
        
        return list(unique_fields)

    async def _validate_unique_fields(
            self,
            db: AsyncSession,
            data: Dict[str, Any],
            exclude_id: str = None
    ) -> None:
        """
        批量校验数据中启用了唯一性校验的字段，不通过则抛出 FormDataException
        
        Args:
            db: 数据库会话
            data: 待写入的主表数据
            exclude_id: 编辑时排除的记录ID
        """
        unique_fields = self._get_unique_check_fields()
        if not unique_fields:
            return

        duplicate_fields = []
        for field_name in unique_fields:
            value = data.get(field_name)
            if value is None or value == '':
                continue
            is_unique = await self.check_unique(db, field_name, str(value), exclude_id)
            if not is_unique:
                duplicate_fields.append(field_name)

        if duplicate_fields:
            field_labels = self._get_field_labels(duplicate_fields)
            messages = [f"'{field_labels.get(f, f)}'" for f in duplicate_fields]
            raise FormDataException(f"以下字段的值已存在: {', '.join(messages)}")

    def _get_field_labels(self, field_names: List[str]) -> Dict[str, str]:
        """获取字段名到标签的映射"""
        labels = {}
        form_config = self._form_config
        items = form_config.get("items", [])

        def traverse(item_list: List[Dict]):
            for item in item_list:
                field = item.get("field", "")
                label = item.get("label", "")
                if field and label and field in field_names:
                    labels[field] = label

                if item.get("columns"):
                    for col in item["columns"]:
                        traverse(col.get("children", []))
                if item.get("items"):
                    for sub_item in item["items"]:
                        traverse(sub_item.get("children", []))
                if item.get("children"):
                    traverse(item["children"])

        traverse(items)
        return labels

    async def check_unique(
            self,
            db: AsyncSession,
            field_name: str,
            value: str,
            exclude_id: str = None
    ) -> bool:
        """
        检查字段值在主表中是否唯一
        
        Args:
            db: 数据库会话
            field_name: 字段名
            value: 字段值
            exclude_id: 排除的记录ID（编辑时排除自身）
        
        Returns:
            True 表示唯一（可用），False 表示已存在
        """
        table = self.form_meta.main_table
        schema = await self._resolve_schema(db, self.form_meta.main_table_schema) or None
        database = self.form_meta.main_table_database or None

        # 验证字段是否存在
        valid_field = await self._validate_sort_field(db, table, schema, database, field_name)
        if not valid_field:
            raise FormDataException(f"字段不存在: {field_name}")

        # 构建查询
        table_name = self.sql_builder.build_table_name(table, schema, database)
        quoted_field = self.sql_builder.quote_identifier(field_name)

        where_clause = f"{quoted_field} = :value"
        params: Dict[str, Any] = {"value": value}

        # 编辑时排除自身
        if exclude_id:
            quoted_id = self.sql_builder.quote_identifier("id")
            where_clause += f" AND {quoted_id} != :exclude_id"
            params["exclude_id"] = exclude_id

        # 检查 is_deleted 字段
        has_is_deleted = await self._check_column_exists(db, table, schema, database, "is_deleted")
        if has_is_deleted:
            where_clause += " AND is_deleted = false"

        sql = f"""
            SELECT COUNT(*) as cnt
            FROM {table_name}
            WHERE {where_clause}
        """
        result = await self._execute_query(db, sql, params)
        count = result[0]["cnt"] if result else 0
        return count == 0

    def _get_field_options(self, field_name: str) -> Optional[List[Dict]]:
        """从表单配置中获取字段的选项列表"""
        form_config = self._form_config
        items = form_config.get("items", [])
        
        def find_field(item_list: List[Dict]) -> Optional[List[Dict]]:
            for item in item_list:
                if item.get("field") == field_name:
                    return item.get("options")
                # 递归查找子项
                for key in ["children", "columns", "items"]:
                    if key in item:
                        sub_items = item[key]
                        if isinstance(sub_items, list):
                            if key == "columns":
                                for col in sub_items:
                                    if "children" in col:
                                        result = find_field(col["children"])
                                        if result is not None:
                                            return result
                            else:
                                result = find_field(sub_items)
                                if result is not None:
                                    return result
            return None
        
        return find_field(items)

    async def get(self, db: AsyncSession, pk: Any) -> Dict[str, Any]:
        """获取单条数据（含子表）"""
        table = self.form_meta.main_table
        schema = await self._resolve_schema(db, self.form_meta.main_table_schema) or None
        database = self.form_meta.main_table_database or None

        # 查询主表
        sql, params = self.sql_builder.build_select(
            table=table,
            schema=schema,
            database=database,
            where={"id": pk}
        )
        rows = await self._execute_query(db, sql, params)

        if not rows:
            raise FormDataException(f"数据不存在: {pk}")

        result = self._serialize_row(rows[0])

        # 查询子表数据
        result["sub_tables"] = {}
        for sub_table in self.sub_tables:
            sub_data = await self._query_sub_table_data(db, sub_table, pk)
            result["sub_tables"][sub_table.table_name] = sub_data

        # 填充虚拟字段的值（基于值关联配置）
        relation_fields = self._get_relation_fields()
        filled = await self._fill_virtual_fields(db, [result], relation_fields, context="form")
        if filled:
            result = filled[0]

        # 应用字段权限过滤
        result = await self.apply_field_permissions(result, db)

        return result

    async def _query_sub_table_data(
            self,
            db: AsyncSession,
            sub_table: FormSubTable,
            main_pk: Any
    ) -> List[Dict[str, Any]]:
        """查询子表数据"""
        sql, params = self.sql_builder.build_select(
            table=sub_table.table_name,
            schema=sub_table.table_schema or None,
            database=sub_table.table_database or None,
            where={sub_table.foreign_key: main_pk}
        )
        rows = await self._execute_query(db, sql, params)
        return [self._serialize_row(row) for row in rows]

    def _is_uuid_like(self, value: str) -> bool:
        """检查值是否像 UUID（用于判断是 ID 还是名称）"""
        if not isinstance(value, str):
            return False
        # UUID 格式: 8-4-4-4-12 或 32位无连字符
        import re
        uuid_pattern = re.compile(
            r'^[0-9a-f]{8}-[0-9a-f]{4}-[0-9a-f]{4}-[0-9a-f]{4}-[0-9a-f]{12}$|^[0-9a-f]{32}$',
            re.IGNORECASE
        )
        return bool(uuid_pattern.match(value))

    def _is_date_string(self, value: str) -> bool:
        """检查值是否为日期字符串格式 YYYY-MM-DD"""
        if not isinstance(value, str):
            return False
        import re
        return bool(re.match(r'^\d{4}-\d{2}-\d{2}$', value))

    def _is_datetime_string(self, value: str) -> bool:
        """检查值是否为日期时间字符串格式 YYYY-MM-DD HH:MM:SS 或 ISO 格式"""
        if not isinstance(value, str):
            return False
        import re
        # 支持多种格式：YYYY-MM-DD HH:MM:SS, YYYY-MM-DDTHH:MM:SS, 带毫秒和时区
        return bool(re.match(r'^\d{4}-\d{2}-\d{2}[T ]\d{2}:\d{2}:\d{2}', value))

    def _parse_datetime_string(self, value: str) -> Optional[datetime]:
        """解析日期时间字符串为 datetime 对象"""
        if not isinstance(value, str):
            return None
        
        # 尝试多种格式
        formats = [
            "%Y-%m-%d %H:%M:%S",
            "%Y-%m-%dT%H:%M:%S",
            "%Y-%m-%dT%H:%M:%S.%f",
            "%Y-%m-%d",
        ]
        
        # 处理带时区的格式
        clean_value = value
        if value.endswith('Z'):
            clean_value = value[:-1]
        elif '+' in value[-6:] or (value.count('-') > 2 and ':' in value[-6:]):
            # 移除时区信息
            clean_value = value.rsplit('+', 1)[0].rsplit('-', 1)[0] if '+' in value[-6:] else value
        
        for fmt in formats:
            try:
                return datetime.strptime(clean_value, fmt)
            except ValueError:
                continue
        
        return None

    @staticmethod
    def _convert_bool_string(value: str):
        """将字符串 'true'/'false' 转换为 Python bool，非布尔字符串返回 None"""
        if value.lower() == 'true':
            return True
        if value.lower() == 'false':
            return False
        return None

    def _convert_date_strings(self, filters: Dict[str, Any]) -> Dict[str, Any]:
        """将过滤器中的日期/日期时间/布尔字符串转换为原生类型（PostgreSQL asyncpg 需要）"""
        if not filters:
            return filters
        
        converted = {}
        for field, value in filters.items():
            if isinstance(value, dict):
                filter_type = value.get("type")
                filter_value = value.get("value")
                
                # 处理布尔字符串
                if filter_type in ("eq",) and isinstance(filter_value, str):
                    bool_val = self._convert_bool_string(filter_value)
                    if bool_val is not None:
                        converted[field] = {"type": filter_type, "value": bool_val}
                        continue

                # 处理 gte, lte, gt, lt, eq 等类型的日期/日期时间值
                # 注意：如果 case_sensitive 明确设为 false，说明是文本比较，不转换日期
                case_sensitive = value.get("case_sensitive", True)
                if filter_type in ("gte", "lte", "gt", "lt", "eq") and isinstance(filter_value, str) and case_sensitive:
                    if self._is_datetime_string(filter_value):
                        # 优先尝试解析为 datetime
                        parsed = self._parse_datetime_string(filter_value)
                        if parsed:
                            converted[field] = {"type": filter_type, "value": parsed}
                        else:
                            converted[field] = value
                    elif self._is_date_string(filter_value):
                        # 对于 datetime 字段使用日期格式查询时，需要特殊处理
                        # lte/lt: 使用当天的 23:59:59.999999（包含当天所有时间）
                        # gte/gt: 使用当天的 00:00:00
                        try:
                            parsed_date = date.fromisoformat(filter_value)
                            if filter_type == "lte":
                                # 小于等于：转换为当天结束时间
                                end_of_day = datetime.combine(parsed_date, datetime.max.time())
                                converted[field] = {"type": filter_type, "value": end_of_day}
                            elif filter_type == "lt":
                                # 小于：转换为当天开始时间（不包含当天）
                                start_of_day = datetime.combine(parsed_date, datetime.min.time())
                                converted[field] = {"type": filter_type, "value": start_of_day}
                            elif filter_type == "gte":
                                # 大于等于：转换为当天开始时间
                                start_of_day = datetime.combine(parsed_date, datetime.min.time())
                                converted[field] = {"type": filter_type, "value": start_of_day}
                            elif filter_type == "gt":
                                # 大于：转换为当天结束时间（不包含当天）
                                end_of_day = datetime.combine(parsed_date, datetime.max.time())
                                converted[field] = {"type": filter_type, "value": end_of_day}
                            elif filter_type == "eq":
                                # 精确匹配日期：转换为当天范围查询（00:00:00 - 23:59:59）
                                # 这样 2026-01-14 可以匹配 2026-01-14 16:16:09
                                start_of_day = datetime.combine(parsed_date, datetime.min.time())
                                end_of_day = datetime.combine(parsed_date, datetime.max.time())
                                converted[field] = {"type": "range", "value": [start_of_day, end_of_day]}
                            else:
                                converted[field] = {"type": filter_type, "value": parsed_date}
                        except ValueError:
                            converted[field] = value
                    else:
                        converted[field] = value
                # 处理 in 类型的日期值列表
                elif filter_type == "in" and isinstance(filter_value, list):
                    converted_values = []
                    for v in filter_value:
                        if isinstance(v, str):
                            if self._is_datetime_string(v):
                                parsed = self._parse_datetime_string(v)
                                converted_values.append(parsed if parsed else v)
                            elif self._is_date_string(v):
                                try:
                                    converted_values.append(date.fromisoformat(v))
                                except ValueError:
                                    converted_values.append(v)
                            else:
                                converted_values.append(v)
                        else:
                            converted_values.append(v)
                    converted[field] = {"type": filter_type, "value": converted_values}
                # 处理 range 类型
                elif filter_type == "range" and isinstance(filter_value, list) and len(filter_value) == 2:
                    converted_values = []
                    for idx, v in enumerate(filter_value):
                        if isinstance(v, str):
                            if self._is_datetime_string(v):
                                parsed = self._parse_datetime_string(v)
                                converted_values.append(parsed if parsed else v)
                            elif self._is_date_string(v):
                                # 对于 range 查询，第一个值（开始）使用当天开始，第二个值（结束）使用当天结束
                                try:
                                    parsed_date = date.fromisoformat(v)
                                    if idx == 0:
                                        # 开始日期：当天 00:00:00
                                        start_of_day = datetime.combine(parsed_date, datetime.min.time())
                                        converted_values.append(start_of_day)
                                    else:
                                        # 结束日期：当天 23:59:59.999999
                                        end_of_day = datetime.combine(parsed_date, datetime.max.time())
                                        converted_values.append(end_of_day)
                                except ValueError:
                                    converted_values.append(v)
                            else:
                                converted_values.append(v)
                        else:
                            converted_values.append(v)
                    converted[field] = {"type": filter_type, "value": converted_values}
                else:
                    converted[field] = value
            elif isinstance(value, str):
                # 简单等值条件：布尔字符串
                bool_val = self._convert_bool_string(value)
                if bool_val is not None:
                    converted[field] = bool_val
                # 简单等值条件的日期/日期时间字符串
                elif self._is_datetime_string(value):
                    parsed = self._parse_datetime_string(value)
                    converted[field] = parsed if parsed else value
                elif self._is_date_string(value):
                    try:
                        converted[field] = date.fromisoformat(value)
                    except ValueError:
                        converted[field] = value
                else:
                    converted[field] = value
            else:
                converted[field] = value
        
        return converted

    async def _convert_relation_name_to_id(self, db: AsyncSession, filters: Dict[str, Any]) -> Dict[str, Any]:
        """将关联字段的名称搜索转换为 ID 搜索，或直接使用 ID 列表"""
        if not filters:
            return filters
        
        # 获取关联字段配置
        relation_fields = self._get_relation_fields()
        
        if not relation_fields:
            return filters
        
        converted_filters = {}
        
        for field, value in filters.items():
            # 检查是否是关联字段
            if field in relation_fields:
                relation_config = relation_fields[field]
                relation_table = relation_config.get("relation_table")
                display_column = relation_config.get("display_column", "name")
                
                if not relation_table:
                    converted_filters[field] = value
                    continue
                
                try:
                    # 检查是否直接传递了 ID 列表（来自选择器组件）
                    if isinstance(value, list):
                        # 检查列表中的值是否都是 UUID 格式
                        if all(self._is_uuid_like(v) for v in value):
                            # 直接使用 ID 列表进行 IN 查询
                            if len(value) == 1:
                                converted_filters[field] = value[0]
                            else:
                                converted_filters[field] = {"type": "in", "value": value}
                            continue
                    
                    # 检查是否是单个 UUID（直接传递的 ID）
                    if isinstance(value, str) and self._is_uuid_like(value):
                        converted_filters[field] = value
                        continue
                    
                    # 根据名称查询关联表获取 ID
                    table_name = self.sql_builder.build_table_name(relation_table, None, None)
                    
                    # 处理不同的过滤类型
                    if isinstance(value, dict) and "type" in value:
                        filter_type = value.get("type")
                        filter_value = value.get("value")

                        # 空格搜索类型是纯文本搜索，不需要转换关联字段
                        if filter_type in ("space_like_and", "space_like_or", "space_eq_and", "space_eq_or"):
                            converted_filters[field] = value
                            continue
                        
                        # 如果 filter_value 是 ID 列表，直接使用
                        if filter_type == "in" and isinstance(filter_value, list):
                            if all(self._is_uuid_like(v) for v in filter_value):
                                converted_filters[field] = value
                                continue
                        
                        if filter_type == "like":
                            # 模糊搜索名称
                            query_sql = f"""
                                SELECT {self.sql_builder.quote_identifier("id")}
                                FROM {table_name}
                                WHERE {self.sql_builder.quote_identifier(display_column)} LIKE :search_value
                                AND {self.sql_builder.quote_identifier("is_deleted")} = false
                            """
                            params = {"search_value": f"%{filter_value}%"}
                        else:
                            # 精确搜索名称
                            case_sensitive = value.get("case_sensitive", True)
                            if self.sql_builder.db_type == "postgresql":
                                if case_sensitive:
                                    where_clause = f"{self.sql_builder.quote_identifier(display_column)} = :search_value"
                                else:
                                    where_clause = f"LOWER({self.sql_builder.quote_identifier(display_column)}) = LOWER(:search_value)"
                            else:
                                if case_sensitive:
                                    where_clause = f"{self.sql_builder.quote_identifier(display_column)} = :search_value"
                                else:
                                    where_clause = f"LOWER({self.sql_builder.quote_identifier(display_column)}) = LOWER(:search_value)"
                            query_sql = f"""
                                SELECT {self.sql_builder.quote_identifier("id")}
                                FROM {table_name}
                                WHERE {where_clause}
                                AND {self.sql_builder.quote_identifier("is_deleted")} = false
                            """
                            params = {"search_value": filter_value}
                    else:
                        # 默认使用模糊搜索（更友好）
                        query_sql = f"""
                            SELECT {self.sql_builder.quote_identifier("id")}
                            FROM {table_name}
                            WHERE {self.sql_builder.quote_identifier(display_column)} LIKE :search_value
                            AND {self.sql_builder.quote_identifier("is_deleted")} = false
                        """
                        params = {"search_value": f"%{value}%"}
                    
                    result = await self._execute_query(db, query_sql, params)
                    
                    if result:
                        # 找到匹配的 ID，使用 ID 进行搜索
                        ids = [row["id"] for row in result]
                        if len(ids) == 1:
                            converted_filters[field] = ids[0]
                        else:
                            # 多个匹配，使用 IN 查询
                            converted_filters[field] = {"type": "in", "value": ids}
                    else:
                        # 没有找到匹配的名称，返回一个不存在的 ID（确保搜索结果为空）
                        converted_filters[field] = "00000000-0000-0000-0000-000000000000"
                        
                except Exception as e:
                    logger.error(f"转换关联字段 {field} 搜索失败: {e}")
                    converted_filters[field] = value
            else:
                # 非关联字段，保持原样
                converted_filters[field] = value
        
        return converted_filters

    def _get_relation_fields(self) -> Dict[str, Dict[str, str]]:
        """从表单配置中提取需要关联查询的字段"""
        relation_fields = {}
        list_config = self._list_config
        columns = list_config.get("columns", [])
        form_config = self._form_config
        
        logger.debug(f"表单配置 form_config 类型: {type(form_config)}")
        logger.debug(f"表单配置内容: {form_config}")
        logger.debug(f"表单配置 items 数量: {len(form_config.get('items', []))}")
        logger.debug(f"列表配置 columns 数量: {len(columns)}")
        
        # 构建字段到组件配置的映射
        field_to_component = {}
        
        def traverse_items(items):
            """递归遍历表单项，构建字段映射"""
            if not items:
                return
            for item in items:
                item_type = item.get("type", "")
                field = item.get("field", "")
                
                # 容器类型，递归处理子项
                if item_type == "grid":
                    for col in item.get("columns", []):
                        traverse_items(col.get("children", []))
                elif item_type in ("collapse", "steps"):
                    for panel in item.get("items", []):
                        traverse_items(panel.get("children", []))
                elif item_type == "tabs":
                    for panel in item.get("children", []):
                        traverse_items(panel.get("children", []))
                elif item_type == "sub-table":
                    traverse_items(item.get("children", []))
                elif field:
                    # 保存字段对应的组件配置
                    field_to_component[field] = item
                    logger.debug(f"找到字段: {field}, 类型: {item_type}")
        
        traverse_items(form_config.get("items", []))
        
        logger.debug(f"构建的字段映射: {list(field_to_component.keys())}")
        
        # 遍历列表配置，查找需要关联查询的字段
        for col in columns:
            field = col.get("field")
            display_field = col.get("displayField")
            
            if not field or not display_field:
                continue
            
            logger.debug(f"检查字段 {field}，displayField: {display_field}")
            
            # 从表单配置中找到对应的组件
            component = field_to_component.get(field)
            if not component:
                # 组件可能在未遍历到的容器中，尝试从列配置的保存信息中恢复
                if col.get("isFormDataSelector") and col.get("formCode"):
                    display_field_name = col.get("displayFieldName", "")
                    form_code = col.get("formCode", "")
                    relation_key = col.get("valueField", "id") or "id"
                    display_column = display_field_name or col.get("labelField", "name") or "name"
                    relation_fields[field] = {
                        "display_field": display_field,
                        "relation_type": "form_data",
                        "form_code": form_code,
                        "relation_key": relation_key,
                        "display_column": display_column,
                    }
                    logger.debug(f"字段 {field} 组件未找到，从列配置恢复: formCode={form_code}, displayColumn={display_column}")
                else:
                    logger.warning(f"字段 {field} 在表单配置中未找到对应组件")
                continue
            
            # 获取组件的 props
            props = component.get("props", {})
            logger.debug(f"字段 {field} 的组件类型: {component.get('type')}, props: {props}")
            
            # 从 props 中提取关联表信息
            relation_table = props.get("relationTable")
            relation_key = props.get("relationKey", "id")
            display_column = props.get("displayColumn", "name")
            
            # 如果 props 中没有配置，尝试根据组件类型和字段名推断
            if not relation_table:
                component_type = component.get("type", "")
                # 根据组件类型推断
                type_mapping = {
                    "department-selector": "core_dept",
                    "dept-selector": "core_dept",
                    "dept-select": "core_dept",
                    "user-selector": "core_user",
                    "user-select": "core_user",
                    "position-selector": "core_post",
                    "post-selector": "core_post",
                    "role-selector": "core_role",
                    "org-selector": "core_dept",
                    "region-selector": "core_region",
                }
                relation_table = type_mapping.get(component_type)
                
                # 如果还是没有，根据字段名推断
                if not relation_table:
                    field_mapping = {
                        "dept_id": "core_dept",
                        "department_id": "core_dept",
                        "user_id": "core_user",
                        "manger_id": "core_user",
                        "manager_id": "core_user",
                        "position_id": "core_post",
                        "positon_id": "core_post",
                    }
                    relation_table = field_mapping.get(field)
            
            if relation_table:
                relation_fields[field] = {
                    "display_field": display_field,
                    "relation_table": relation_table,
                    "relation_key": relation_key,
                    "display_column": display_column,
                }
                logger.debug(f"字段 {field} 配置了关联表: {relation_table}")
            else:
                # 检查是否为表单数据选择器（select 或 table-selector 组件，数据源类型为 formData）
                # 或者是新的 form-selector 组件
                component_type = component.get("type", "")
                
                # form-selector 组件的配置存储在 formSelectorConfig 中
                form_selector_config = component.get("formSelectorConfig") or {}
                
                # 数据源配置存储在 dataSource 中，而不是 props 中
                data_source = component.get("dataSource") or {}
                data_source_type = data_source.get("type", "") if data_source else ""
                if not data_source_type:
                    data_source_type = props.get("dataSourceType", "")
                
                # 获取 formCode：优先从 formSelectorConfig 获取，其次从 dataSource 获取
                form_code = form_selector_config.get("formCode", "") or data_source.get("formCode", "") or props.get("formCode", "")
                
                display_field_name = col.get("displayFieldName", "")  # 从列表配置中获取要显示的字段名
                
                # form-selector 组件直接处理
                if component_type == "form-selector" and form_code:
                    relation_key = form_selector_config.get("valueField", "id") or "id"
                    label_field = form_selector_config.get("labelField", "name") or "name"
                    relation_fields[field] = {
                        "display_field": display_field,
                        "relation_type": "form_data",
                        "form_code": form_code,
                        "relation_key": relation_key,
                        "display_column": display_field_name or label_field,
                    }
                    logger.debug(f"字段 {field} 配置了表单选择器关联: formCode={form_code}, relationKey={relation_key}, displayColumn={display_field_name or label_field}")
                elif component_type in ["select", "tree-select", "radio", "checkbox", "cascader", "table-selector"] and data_source_type == "formData" and form_code:
                    relation_key = data_source.get("formValueField") or data_source.get("valueField", "id") or "id"
                    label_field = data_source.get("formLabelField") or data_source.get("labelField", "name") or "name"
                    relation_fields[field] = {
                        "display_field": display_field,
                        "relation_type": "form_data",
                        "form_code": form_code,
                        "relation_key": relation_key,
                        "display_column": display_field_name or label_field,
                    }
                    logger.debug(f"字段 {field} 配置了表单数据关联: formCode={form_code}, relationKey={relation_key}, displayColumn={display_field_name or label_field}")
                else:
                    logger.warning(f"字段 {field} 未配置 relationTable 且无法自动推断")
        
        # 添加系统字段的关联配置（这些字段不在表单配置中，但需要关联查询）
        system_field_mapping = {
            "sys_creator_id": {"relation_table": "core_user", "display_column": "name"},
            "sys_modifier_id": {"relation_table": "core_user", "display_column": "name"},
            "sys_dept_id": {"relation_table": "core_dept", "display_column": "name"},
        }
        
        for col in columns:
            field = col.get("field")
            display_field = col.get("displayField")
            
            if field in system_field_mapping and display_field:
                config = system_field_mapping[field]
                relation_fields[field] = {
                    "display_field": display_field,
                    "relation_table": config["relation_table"],
                    "relation_key": "id",
                    "display_column": config["display_column"],
                }
                logger.debug(f"系统字段 {field} 配置了关联表: {config['relation_table']}")

        # 识别没有 displayField 的数据源字段（字典 / 表单数据），原地替换 value 为 label
        column_fields = {col.get("field") for col in columns if col.get("field")}
        for field, component in field_to_component.items():
            if field in relation_fields or field not in column_fields:
                continue
            data_source = component.get("dataSource") or {}
            ds_type = data_source.get("type", "")

            if ds_type == "dict" and data_source.get("dictCode"):
                relation_fields[field] = {
                    "relation_type": "dict",
                    "dict_code": data_source["dictCode"],
                }
                logger.debug(f"字段 {field} 配置了字典数据源: dictCode={data_source['dictCode']}")
            elif ds_type == "formData" and data_source.get("formCode"):
                component_type = component.get("type", "")
                if component_type in ["select", "tree-select", "radio", "checkbox", "cascader", "table-selector"]:
                    form_code = data_source["formCode"]
                    relation_key = data_source.get("formValueField") or data_source.get("valueField", "id") or "id"
                    label_field = data_source.get("formLabelField") or data_source.get("labelField", "name") or "name"
                    display_field = f"{field}_name"
                    relation_fields[field] = {
                        "display_field": display_field,
                        "relation_type": "form_data",
                        "form_code": form_code,
                        "relation_key": relation_key,
                        "display_column": label_field,
                    }
                    logger.debug(f"字段 {field} 无 displayField，自动补充表单数据关联: formCode={form_code}, displayField={display_field}")

        return relation_fields

    async def _fill_relation_display_names(
            self,
            db: AsyncSession,
            items: List[Dict[str, Any]],
            relation_fields: Dict[str, Dict[str, str]]
    ) -> List[Dict[str, Any]]:
        """填充关联字段的显示名称"""
        for field, config in relation_fields.items():
            relation_type = config.get("relation_type", "")
            relation_table = config.get("relation_table", "")
            
            # 表单数据选择器特殊处理
            if relation_type == "form_data":
                try:
                    await self._fill_form_data_display_field(db, items, field, config)
                except Exception as e:
                    logger.warning(f"填充表单数据字段 {field} 失败: {str(e)}")
                    try:
                        await db.rollback()
                    except Exception:
                        pass
                continue

            # 字典数据源：批量查询 dict_item，原地替换 value 为 label
            if relation_type == "dict":
                dict_code = config.get("dict_code")
                if not dict_code:
                    continue
                try:
                    values = set()
                    for item in items:
                        v = item.get(field)
                        if v is not None:
                            is_list, parsed = self._parse_list_value(v)
                            if is_list:
                                values.update(str(x) for x in parsed if x is not None)
                            else:
                                values.add(str(v))
                    if not values:
                        continue
                    value_to_label = await self._query_dict_labels(db, dict_code, list(values))
                    for item in items:
                        v = item.get(field)
                        if v is not None:
                            is_list, parsed = self._parse_list_value(v)
                            if is_list:
                                item[field] = [value_to_label.get(str(x), x) for x in parsed]
                            else:
                                item[field] = value_to_label.get(str(v), v)
                    logger.debug(f"字典字段 {field} 翻译完成: dictCode={dict_code}, 映射数={len(value_to_label)}")
                except Exception as e:
                    logger.warning(f"填充字典字段 {field} 失败: {str(e)}")
                continue

            # 如果没有关联表，跳过
            if not relation_table:
                logger.warning(f"字段 {field} 没有配置关联表，跳过")
                continue
            
            # 省市区组件特殊处理
            if relation_table == "core_region":
                try:
                    # 收集所有需要查询的 code（省市区组件存储的是 code 数组）
                    codes = set()
                    for item in items:
                        value = item.get(field)
                        if value and isinstance(value, list):
                            codes.update(str(v) for v in value if v)
                    
                    if not codes:
                        continue
                    
                    # 批量查询省市区名称
                    code_to_name = await self._query_region_names(db, list(codes))
                    display_field = config["display_field"]
                    
                    # 填充显示名称（按顺序拼接，用 / 分隔）
                    for item in items:
                        value = item.get(field)
                        if value and isinstance(value, list):
                            names = [code_to_name.get(code, code) for code in value]
                            item[display_field] = " / ".join(names)
                except Exception as e:
                    logger.warning(f"填充省市区字段 {field} 失败: {str(e)}")
                    continue
            elif relation_table:
                # 其他关联组件的处理逻辑（用户、部门等）
                # 收集所有需要查询的 ID
                ids = set()
                for item in items:
                    value = item.get(field)
                    if value:
                        if isinstance(value, list):
                            ids.update(str(v) for v in value if v)
                        else:
                            ids.add(str(value))
                
                if not ids:
                    continue
                
                logger.info(f"准备查询关联字段 {field}，关联表: {relation_table}，IDs: {ids}")
                
                try:
                    # 批量查询关联数据
                    relation_data = await self._query_relation_data(
                        db,
                        config["relation_table"],
                        config["relation_key"],
                        config["display_column"],
                        list(ids)
                    )
                    
                    # 构建 ID 到名称的映射
                    id_to_name = {str(row["id"]): row["name"] for row in relation_data}
                    display_field = config["display_field"]
                    
                    logger.info(f"字段 {field} 的 ID 到名称映射: {id_to_name}")
                    
                    # 填充显示名称
                    for item in items:
                        value = item.get(field)
                        if value:
                            if isinstance(value, list):
                                # 多选情况
                                names = [id_to_name.get(str(v), str(v)) for v in value if v]
                                item[display_field] = ", ".join(names)
                            else:
                                # 单选情况
                                item[display_field] = id_to_name.get(str(value), str(value))
                except Exception as e:
                    logger.warning(f"填充关联字段 {field} 失败: {str(e)}")
                    # 回滚事务，避免影响后续查询
                    try:
                        await db.rollback()
                    except Exception:
                        pass
                    # 填充失败不影响主流程，继续处理其他字段
                    continue
        
        return items

    @staticmethod
    def _parse_list_value(v: Any) -> Tuple[bool, list]:
        """将字段值解析为列表。
        级联/多选组件的值可能是 Python list（json/jsonb 列）、JSON 字符串（varchar/text 列），
        或历史数据中 Python str(list) 产生的单引号格式（如 "['a', 'b']"）。
        返回 (is_list, parsed_list)；非列表值返回 (False, [])。
        """
        if isinstance(v, list):
            return True, v
        if isinstance(v, str) and v.startswith("["):
            try:
                parsed = json.loads(v)
                if isinstance(parsed, list):
                    return True, parsed
            except (json.JSONDecodeError, ValueError):
                pass
            # 兼容历史数据：str(list) 产生的单引号格式，如 "['a', 'b']"
            try:
                import ast
                parsed = ast.literal_eval(v)
                if isinstance(parsed, list):
                    return True, parsed
            except (ValueError, SyntaxError):
                pass
        return False, []

    async def _query_dict_labels(self, db: AsyncSession, dict_code: str, values: List[str]) -> Dict[str, str]:
        """批量查询字典项的 value->label 映射"""
        from core.dict.model import Dict as DictModel
        from core.dict_item.model import DictItem

        stmt = (
            select(DictItem.value, DictItem.label)
            .join(DictModel, DictItem.dict_id == DictModel.id)
            .where(
                DictModel.code == dict_code,
                DictItem.value.in_(values),
                DictItem.is_deleted == False,
            )
        )
        result = await db.execute(stmt)
        return {row.value: row.label for row in result}

    async def _query_form_data_labels(
            self,
            db: AsyncSession,
            field: str,
            config: Dict[str, Any],
            items: List[Dict[str, Any]]
    ) -> Dict[str, str]:
        """查询关联表单数据，返回 value->label 映射（用于无 displayField 的表单数据源字段）"""
        form_code = config.get("form_code")
        relation_key = config.get("relation_key", "id")
        display_column = config.get("display_column", "name")

        # 收集所有不重复的 value（兼容 JSON 字符串数组）
        ids = set()
        for item in items:
            v = item.get(field)
            if v is not None:
                is_list, parsed = self._parse_list_value(v)
                if is_list:
                    ids.update(str(x) for x in parsed if x is not None)
                else:
                    ids.add(str(v))

        if not ids or not form_code:
            return {}

        from online_dev.form_manager.model import FormMeta as RelatedFormMeta

        stmt = select(RelatedFormMeta).where(
            RelatedFormMeta.code == form_code,
            RelatedFormMeta.is_deleted == False
        )
        result = await db.execute(stmt)
        related_form = result.scalar_one_or_none()

        if not related_form:
            logger.warning(f"关联表单 {form_code} 不存在")
            return {}

        table = related_form.main_table
        schema = await self._resolve_schema(db, related_form.main_table_schema) or None
        database = related_form.main_table_database or None

        table_name = self.sql_builder.build_table_name(table, schema, database)
        key_col = self.sql_builder.quote_identifier(relation_key)
        display_col = self.sql_builder.quote_identifier(display_column)

        # 检查关联表是否有 is_deleted 列
        has_is_deleted = await self._check_column_exists(db, table,
                                                         related_form.main_table_schema,
                                                         related_form.main_table_database,
                                                         "is_deleted")

        ids_list = list(ids)
        placeholders = ", ".join([f":{i}" for i in range(len(ids_list))])
        is_deleted_clause = "\n                  AND is_deleted = false" if has_is_deleted else ""
        sql = f"""
            SELECT {key_col} as id, {display_col} as name
            FROM {table_name}
            WHERE {key_col} IN ({placeholders}){is_deleted_clause}
        """
        params = {str(i): ids_list[i] for i in range(len(ids_list))}
        rows = await self._execute_query(db, sql, params)

        return {str(row["id"]): str(row["name"]) for row in rows}

    async def _fill_virtual_fields(
            self,
            db: AsyncSession,
            items: List[Dict[str, Any]],
            relation_fields: Dict[str, Dict[str, str]],
            context: str = "list"
    ) -> List[Dict[str, Any]]:
        """填充虚拟字段的值
        
        虚拟字段通过值关联配置（isVirtualField=true），从源字段的关联数据中
        提取指定属性填充到结果中，不对应数据库列。
        
        Args:
            db: 数据库会话
            items: 数据列表
            relation_fields: 已识别的关联字段配置（来自 _get_relation_fields）
            context: 调用上下文，'list' 列表接口（受 showVirtualValue 开关控制），'form' 表单详情（始终填充）
        """
        if not items:
            return items
        
        # 从 form_config 中提取虚拟字段配置
        virtual_fields = self._get_virtual_fields()
        if not virtual_fields:
            return items
        
        # 构建 list_config 中虚拟字段的 showVirtualValue 映射
        list_config = self._list_config
        columns = list_config.get("columns", [])
        col_show_map = {}
        for col in columns:
            if col.get("isVirtualField"):
                col_show_map[col.get("field", "")] = col.get("showVirtualValue", True)
        
        logger.info(f"识别到的虚拟字段: {virtual_fields}")
        
        for vf in virtual_fields:
            vf_field = vf["field"]  # 虚拟字段名
            source_field = vf["valueSourceField"]  # 源字段名
            display_field = vf["valueDisplayField"]  # 要提取的属性名
            
            # 仅在列表上下文中检查 showVirtualValue 开关，表单详情始终填充
            if context == "list" and not col_show_map.get(vf_field, True):
                logger.info(f"虚拟字段 {vf_field} 的显示关联值已关闭，跳过")
                continue
            
            # 从 relation_fields 中获取源字段的关联配置
            source_config = relation_fields.get(source_field)
            
            # 如果源字段不在 relation_fields 中（可能未添加到列表列），从 form_config 推断
            if not source_config:
                source_config = self._infer_relation_config(source_field)
            
            if not source_config:
                logger.warning(f"虚拟字段 {vf_field} 的源字段 {source_field} 没有关联配置，跳过")
                continue
            
            relation_type = source_config.get("relation_type", "")
            relation_table = source_config.get("relation_table", "")
            relation_key = source_config.get("relation_key", "id")
            
            # 省市区组件不支持虚拟字段
            if relation_table == "core_region":
                logger.warning(f"虚拟字段 {vf_field} 的源字段 {source_field} 是省市区组件，不支持虚拟字段")
                continue
            
            # 收集源字段的所有 ID
            ids = set()
            for item in items:
                value = item.get(source_field)
                if value:
                    if isinstance(value, list):
                        ids.update(str(v) for v in value if v)
                    else:
                        ids.add(str(value))
            
            if not ids:
                continue
            
            try:
                id_to_value = {}
                
                if relation_type == "form_data":
                    # 表单数据选择器：查询关联表单的数据表
                    form_code = source_config.get("form_code")
                    if not form_code:
                        continue
                    
                    from online_dev.form_manager.model import FormMeta as RelatedFormMeta
                    stmt = select(RelatedFormMeta).where(
                        RelatedFormMeta.code == form_code,
                        RelatedFormMeta.is_deleted == False
                    )
                    result = await db.execute(stmt)
                    related_form = result.scalar_one_or_none()
                    
                    if not related_form:
                        logger.warning(f"虚拟字段 {vf_field} 关联表单 {form_code} 不存在")
                        continue
                    
                    table = related_form.main_table
                    schema = await self._resolve_schema(db, related_form.main_table_schema) or None
                    database = related_form.main_table_database or None
                    table_name = self.sql_builder.build_table_name(table, schema, database)
                    key_col = self.sql_builder.quote_identifier(relation_key)
                    val_col = self.sql_builder.quote_identifier(display_field)
                    
                    # 检查关联表是否有 is_deleted 列
                    has_is_deleted = False
                    related_form_config = related_form.form_config or {}
                    for tc in related_form_config.get("tableConfigs", []):
                        if tc.get("type") == "main":
                            for f in tc.get("fields", []):
                                if f.get("name") == "is_deleted":
                                    has_is_deleted = True
                                    break
                            break
                    
                    placeholders = ", ".join([f":{i}" for i in range(len(ids))])
                    ids_list = list(ids)
                    is_deleted_clause = "\n                          AND is_deleted = false" if has_is_deleted else ""
                    sql = f"""
                        SELECT {key_col} as id, {val_col} as val
                        FROM {table_name}
                        WHERE {key_col} IN ({placeholders}){is_deleted_clause}
                    """
                    params = {str(i): ids_list[i] for i in range(len(ids_list))}
                    rows = await self._execute_query(db, sql, params)
                    id_to_value = {str(row["id"]): row["val"] for row in rows}
                    
                else:
                    # 普通关联表（用户、部门等）
                    if not relation_table:
                        continue
                    
                    table_name = self.sql_builder.build_table_name(relation_table, schema="public")
                    key_col = self.sql_builder.quote_identifier(relation_key)
                    val_col = self.sql_builder.quote_identifier(display_field)
                    
                    placeholders = ", ".join([f":{i}" for i in range(len(ids))])
                    ids_list = list(ids)
                    sql = f"""
                        SELECT {key_col} as id, {val_col} as val
                        FROM {table_name}
                        WHERE {key_col} IN ({placeholders})
                          AND {self.sql_builder.quote_identifier("is_deleted")} = false
                    """
                    params = {str(i): ids_list[i] for i in range(len(ids_list))}
                    rows = await self._execute_query(db, sql, params)
                    id_to_value = {str(row["id"]): row["val"] for row in rows}
                
                # 填充虚拟字段值
                for item in items:
                    source_value = item.get(source_field)
                    if source_value:
                        if isinstance(source_value, list):
                            vals = [id_to_value.get(str(v), "") for v in source_value if v]
                            item[vf_field] = ", ".join(str(v) for v in vals if v)
                        else:
                            item[vf_field] = id_to_value.get(str(source_value), "")
                    else:
                        item[vf_field] = ""
                        
            except Exception as e:
                logger.warning(f"填充虚拟字段 {vf_field} 失败: {str(e)}")
                try:
                    await db.rollback()
                except Exception:
                    pass
                continue
        
        return items

    def _get_virtual_fields(self) -> List[Dict[str, str]]:
        """从 form_config 中提取虚拟字段配置
        
        Returns:
            虚拟字段配置列表，每项包含:
            - field: 虚拟字段名
            - valueSourceField: 源字段名
            - valueDisplayField: 要提取的属性名
        """
        virtual_fields = []
        form_config = self._form_config
        
        def traverse(items):
            if not items:
                return
            for item in items:
                item_type = item.get("type", "")
                field = item.get("field", "")
                props = item.get("props", {})
                
                # 检查是否为虚拟字段
                if (field and props.get("isVirtualField") and 
                        props.get("enableValueLink") and 
                        props.get("valueSourceField") and 
                        props.get("valueDisplayField")):
                    virtual_fields.append({
                        "field": field,
                        "valueSourceField": props["valueSourceField"],
                        "valueDisplayField": props["valueDisplayField"],
                    })
                
                # 递归处理容器
                if item_type == "grid":
                    for col in item.get("columns", []):
                        traverse(col.get("children", []))
                elif item_type in ("collapse", "steps"):
                    for panel in item.get("items", []):
                        traverse(panel.get("children", []))
                elif item_type == "tabs":
                    for panel in item.get("children", []):
                        traverse(panel.get("children", []))
                elif item_type == "sub-table":
                    traverse(item.get("children", []))
        
        traverse(form_config.get("items", []))
        return virtual_fields

    def _infer_relation_config(self, source_field: str) -> Optional[Dict[str, str]]:
        """从 form_config 中推断源字段的关联表配置
        
        当源字段未出现在 list_config.columns 中时，需要从 form_config 的组件类型推断关联信息。
        
        Args:
            source_field: 源字段名
            
        Returns:
            关联配置字典，包含 relation_table, relation_key, display_column 等，或 None
        """
        form_config = self._form_config
        
        # 递归查找源字段的组件配置
        def find_component(items):
            if not items:
                return None
            for item in items:
                item_type = item.get("type", "")
                field = item.get("field", "")
                if field == source_field:
                    return item
                if item_type == "grid":
                    for col in item.get("columns", []):
                        found = find_component(col.get("children", []))
                        if found:
                            return found
                elif item_type in ("collapse", "steps"):
                    for panel in item.get("items", []):
                        found = find_component(panel.get("children", []))
                        if found:
                            return found
                elif item_type == "tabs":
                    for panel in item.get("children", []):
                        found = find_component(panel.get("children", []))
                        if found:
                            return found
                elif item_type == "sub-table":
                    found = find_component(item.get("children", []))
                    if found:
                        return found
            return None
        
        component = find_component(form_config.get("items", []))
        if not component:
            return None
        
        component_type = component.get("type", "")
        props = component.get("props", {})
        
        # 根据组件类型推断关联表
        type_mapping = {
            "department-selector": "core_dept",
            "dept-selector": "core_dept",
            "dept-select": "core_dept",
            "user-selector": "core_user",
            "user-select": "core_user",
            "position-selector": "core_post",
            "post-selector": "core_post",
            "role-selector": "core_role",
            "org-selector": "core_dept",
            "region-selector": "core_region",
        }
        
        relation_table = props.get("relationTable") or type_mapping.get(component_type)
        
        if relation_table:
            return {
                "relation_table": relation_table,
                "relation_key": props.get("relationKey", "id"),
                "display_column": "name",
            }
        
        # form-selector 组件
        if component_type == "form-selector":
            form_selector_config = component.get("formSelectorConfig") or {}
            form_code = form_selector_config.get("formCode", "") or props.get("formCode", "")
            if form_code:
                return {
                    "relation_type": "form_data",
                    "form_code": form_code,
                    "relation_key": form_selector_config.get("valueField", "id") or "id",
                    "display_column": form_selector_config.get("labelField", "name") or "name",
                }
        
        # select/table-selector 组件，数据源为 formData
        data_source = component.get("dataSource") or {}
        data_source_type = data_source.get("type", "") or props.get("dataSourceType", "")
        form_code = data_source.get("formCode", "") or props.get("formCode", "")
        
        if component_type in ["select", "tree-select", "radio", "checkbox", "cascader", "table-selector"] and data_source_type == "formData" and form_code:
            return {
                "relation_type": "form_data",
                "form_code": form_code,
                "relation_key": data_source.get("formValueField") or data_source.get("valueField", "id") or "id",
                "display_column": data_source.get("formLabelField") or data_source.get("labelField", "name") or "name",
            }
        
        # 根据字段名推断
        field_mapping = {
            "dept_id": "core_dept",
            "department_id": "core_dept",
            "user_id": "core_user",
            "manger_id": "core_user",
            "manager_id": "core_user",
            "position_id": "core_post",
            "positon_id": "core_post",
        }
        relation_table = field_mapping.get(source_field)
        if relation_table:
            return {
                "relation_table": relation_table,
                "relation_key": "id",
                "display_column": "name",
            }
        
        return None

    async def _fill_form_data_display_field(
            self,
            db: AsyncSession,
            items: List[Dict[str, Any]],
            field: str,
            config: Dict[str, Any]
    ) -> None:
        """填充表单数据选择器的显示字段
        
        Args:
            db: 数据库会话
            items: 数据列表
            field: ID 字段名（如 customer_id）
            config: 配置信息，包含 form_code, display_field, display_column, relation_key
        """
        form_code = config.get("form_code")
        display_field = config.get("display_field")
        display_column = config.get("display_column", "name")
        relation_key = config.get("relation_key", "id")
        
        if not form_code or not display_field:
            return
        
        # 收集所有需要查询的 ID
        ids = set()
        for item in items:
            value = item.get(field)
            if value:
                if isinstance(value, list):
                    ids.update(str(v) for v in value if v)
                else:
                    ids.add(str(value))
        
        if not ids:
            return
        
        try:
            # 获取关联表单的元数据
            from online_dev.form_manager.model import FormMeta as RelatedFormMeta
            from sqlalchemy import select
            
            stmt = select(RelatedFormMeta).where(
                RelatedFormMeta.code == form_code,
                RelatedFormMeta.is_deleted == False
            )
            result = await db.execute(stmt)
            related_form = result.scalar_one_or_none()
            
            if not related_form:
                logger.warning(f"关联表单 {form_code} 不存在")
                return
            
            # 获取关联表单的表信息
            table = related_form.main_table
            schema = await self._resolve_schema(db, related_form.main_table_schema) or None
            database = related_form.main_table_database or None
            
            # 构建查询
            table_name = self.sql_builder.build_table_name(table, schema, database)
            key_col = self.sql_builder.quote_identifier(relation_key)
            display_col = self.sql_builder.quote_identifier(display_column)
            
            # 检查关联表是否有 is_deleted 列
            has_is_deleted = False
            related_form_config = related_form.form_config or {}
            for tc in related_form_config.get("tableConfigs", []):
                if tc.get("type") == "main":
                    for f in tc.get("fields", []):
                        if f.get("name") == "is_deleted":
                            has_is_deleted = True
                            break
                    break
            
            # 使用参数化查询
            placeholders = ", ".join([f":{i}" for i in range(len(ids))])
            ids_list = list(ids)
            
            is_deleted_clause = "\n                  AND is_deleted = false" if has_is_deleted else ""
            sql = f"""
                SELECT {key_col} as id, {display_col} as name
                FROM {table_name}
                WHERE {key_col} IN ({placeholders}){is_deleted_clause}
            """
            
            # 构建参数字典
            params = {str(i): ids_list[i] for i in range(len(ids_list))}
            
            rows = await self._execute_query(db, sql, params)
            logger.debug(f"查询表单数据 {form_code}.{table}，找到 {len(rows)} 条记录")
            
            # 构建 ID 到名称的映射
            id_to_name = {str(row["id"]): row["name"] for row in rows}
            
            # 填充显示名称
            for item in items:
                value = item.get(field)
                if value:
                    if isinstance(value, list):
                        # 多选情况
                        names = [id_to_name.get(str(v), str(v)) for v in value if v]
                        item[display_field] = ", ".join(names)
                    else:
                        # 单选情况
                        item[display_field] = id_to_name.get(str(value), str(value))
                        
        except Exception as e:
            logger.error(f"填充表单数据字段 {field} 失败: {str(e)}")
            raise

    async def _query_region_names(
            self,
            db: AsyncSession,
            codes: List[str]
    ) -> Dict[str, str]:
        """批量查询省市区名称
        
        省市区数据分散在 5 个表中：
        - core_province: 2位代码（如 "14"）
        - core_city: 4位代码（如 "1404"）
        - core_area: 6位代码（如 "140406"）
        - core_street: 9位代码（如 "140406001"）
        - core_village: 12位代码（如 "140406001001"）
        
        Args:
            db: 数据库会话
            codes: 行政区划代码列表
            
        Returns:
            code 到 name 的映射字典
        """
        if not codes:
            return {}
        
        result = {}
        
        try:
            # 按代码长度分组
            codes_by_length = {}
            for code in codes:
                length = len(str(code))
                if length not in codes_by_length:
                    codes_by_length[length] = []
                codes_by_length[length].append(str(code))
            
            # 定义表名映射（代码长度 -> 表名）
            table_mapping = {
                2: "core_province",
                4: "core_city",
                6: "core_area",
                9: "core_street",
                12: "core_village",
            }
            
            # 分别查询每个表
            for length, code_list in codes_by_length.items():
                table_name = table_mapping.get(length)
                if not table_name:
                    logger.warning(f"未知的代码长度: {length}，代码: {code_list}")
                    continue
                
                # 构建表名（使用 public schema）
                full_table_name = self.sql_builder.build_table_name(table_name, schema="public")
                code_col = self.sql_builder.quote_identifier("code")
                name_col = self.sql_builder.quote_identifier("name")
                
                # 使用参数化查询
                placeholders = ", ".join([f":{i}" for i in range(len(code_list))])
                
                sql = f"""
                    SELECT {code_col} as code, {name_col} as name
                    FROM {full_table_name}
                    WHERE {code_col} IN ({placeholders})
                """
                
                # 构建参数字典
                params = {str(i): code_list[i] for i in range(len(code_list))}
                
                rows = await self._execute_query(db, sql, params)
                logger.info(f"查询 {table_name}，找到 {len(rows)} 条记录")
                
                # 添加到结果映射
                for row in rows:
                    result[str(row["code"])] = row["name"]
            
            return result
        except Exception as e:
            logger.error(f"查询省市区名称失败: {str(e)}")
            return {}

    async def _query_relation_data(
            self,
            db: AsyncSession,
            table: str,
            key_column: str,
            display_column: str,
            ids: List[str]
    ) -> List[Dict[str, Any]]:
        """批量查询关联表数据"""
        if not ids:
            return []
        
        # 构建完整的表名（包含 schema）
        # 默认使用 public schema
        table_name = self.sql_builder.build_table_name(table, schema="public")
        key_col = self.sql_builder.quote_identifier(key_column)
        display_col = self.sql_builder.quote_identifier(display_column)
        
        # 使用参数化查询防止 SQL 注入
        placeholders = ", ".join([f":{i}" for i in range(len(ids))])
        
        sql = f"""
            SELECT {key_col} as id, {display_col} as name
            FROM {table_name}
            WHERE {key_col} IN ({placeholders})
            AND {self.sql_builder.quote_identifier("is_deleted")} = false
        """
        
        # 构建参数字典
        params = {str(i): ids[i] for i in range(len(ids))}
        
        try:
            logger.info(f"查询关联表 {table}，SQL: {sql}, params: {params}")
            rows = await self._execute_query(db, sql, params)
            logger.info(f"查询关联表 {table}，找到 {len(rows)} 条记录，结果: {rows}")
            return [dict(row) for row in rows]
        except Exception as e:
            logger.error(f"查询关联表 {table} 失败: {str(e)}, SQL: {sql}, params: {params}")
            return []

    # ============ 写入操作 ============

    def _convert_data_types(self, data: Dict[str, Any]) -> Dict[str, Any]:
        """转换数据类型，将字符串日期时间转换为 Python datetime 对象，处理空字符串"""
        converted = {}
        
        # 获取字段类型映射（用于判断数字类型字段）
        field_types = self._get_field_types()
        
        # 调试日志
        import logging
        logger = logging.getLogger(__name__)
        # logger.info(f"字段类型映射: {field_types}")
        # logger.info(f"待转换数据: {data}")
        
        for key, value in data.items():
            if value is None:
                converted[key] = value
                continue
            
            # 处理空字符串
            if isinstance(value, str) and value.strip() == '':
                # 获取字段类型
                field_type = field_types.get(key, '').lower()
                # 数字类型字段的空字符串转为 None
                if field_type in ('int', 'integer', 'bigint', 'smallint', 'decimal', 'numeric', 'float', 'double', 'real'):
                    converted[key] = None
                    continue
                # 其他类型保持空字符串或转为 None
                converted[key] = None
                continue
            
            # 获取字段类型
            field_type = field_types.get(key, '').lower()
            
            # 处理非字符串类型到字符串的转换
            # 支持 PostgreSQL 的类型名：character varying, character, text 等
            is_string_type = any(t in field_type for t in ['varchar', 'text', 'char', 'string'])
            if not isinstance(value, str) and is_string_type:
                if isinstance(value, datetime):
                    if value.tzinfo is not None:
                        value = value.astimezone(APP_TIMEZONE).replace(tzinfo=None)
                    converted[key] = value.isoformat()
                    logger.info(f"字段 {key} 从 datetime 转换为 str: {value} -> {converted[key]}")
                elif isinstance(value, (list, dict)):
                    converted[key] = json.dumps(value, ensure_ascii=False)
                    logger.info(f"字段 {key} 从 {type(value).__name__} 序列化为 JSON: {converted[key]}")
                else:
                    converted[key] = str(value)
                    logger.info(f"字段 {key} 从 {type(value).__name__} 转换为 str: {value} -> {str(value)}")
                continue
            
            # 处理已经是 datetime 对象的情况
            if isinstance(value, datetime):
                # 如果是 offset-aware datetime，先转换为配置的时区再移除时区信息
                if value.tzinfo is not None:
                    value = value.astimezone(APP_TIMEZONE).replace(tzinfo=None)
                
                # 检查字段类型是否是 timestamp/datetime 类型
                is_datetime_type = any(t in field_type for t in ('date', 'datetime', 'timestamp', 'time'))
                if is_datetime_type:
                    # 保持为 datetime 对象
                    converted[key] = value
                else:
                    # 字段类型未知或是字符串类型，转换为 ISO 格式字符串
                    converted[key] = value.isoformat()
                    logger.info(f"字段 {key} 从 datetime 转换为 str (字段类型: {field_type}): {value} -> {converted[key]}")
                continue
            
            # 转换字符串类型的值
            if isinstance(value, str):
                
                # 1. 尝试转换布尔类型
                if field_type in ('bool', 'boolean'):
                    # 支持多种布尔值表示
                    value_lower = value.lower().strip()
                    if value_lower in ('true', '1', 'yes', 'y', 't', '是', '真'):
                        converted[key] = True
                        continue
                    elif value_lower in ('false', '0', 'no', 'n', 'f', '否', '假'):
                        converted[key] = False
                        continue
                    # 其他值保持原样
                
                # 2. 尝试转换整数类型
                if field_type in ('int', 'integer', 'bigint', 'smallint'):
                    try:
                        # 先去除空格，支持 "25 " 这样的输入
                        converted[key] = int(value.strip())
                        continue
                    except (ValueError, TypeError):
                        pass
                
                # 3. 尝试转换浮点数类型
                elif field_type in ('decimal', 'numeric', 'float', 'double', 'real'):
                    try:
                        converted[key] = float(value.strip())
                        continue
                    except (ValueError, TypeError):
                        pass
                
                # 4. 尝试解析日期时间（支持 'timestamp without time zone' 等完整类型名）
                is_datetime_type = any(t in field_type for t in ('date', 'datetime', 'timestamp', 'time'))
                if is_datetime_type:
                    try:
                        from datetime import time as time_type
                        
                        # 时间格式
                        if field_type == 'time':
                            for fmt in ["%H:%M:%S", "%H:%M:%S.%f", "%H:%M"]:
                                try:
                                    t = datetime.strptime(value, fmt).time()
                                    converted[key] = t
                                    break
                                except ValueError:
                                    continue
                            else:
                                converted[key] = value
                            continue
                        
                        # 先尝试使用 fromisoformat 解析（支持更多 ISO 格式）
                        parsed_dt = None
                        try:
                            # 处理 Z 后缀（UTC 时区标识）
                            iso_value = value.replace('Z', '+00:00') if value.endswith('Z') else value
                            parsed_dt = datetime.fromisoformat(iso_value)
                        except ValueError:
                            # fromisoformat 失败，尝试其他格式
                            for fmt in [
                                "%Y-%m-%d %H:%M:%S",
                                "%Y-%m-%d %H:%M:%S.%f",
                                "%Y-%m-%dT%H:%M:%S",
                                "%Y-%m-%dT%H:%M:%S.%f",
                                "%Y-%m-%d",
                            ]:
                                try:
                                    parsed_dt = datetime.strptime(value, fmt)
                                    break
                                except ValueError:
                                    continue
                        
                        if parsed_dt is not None:
                            # 如果是日期格式（没有时间部分），只保留日期
                            if len(value) == 10 and '-' in value:
                                converted[key] = parsed_dt.date()
                            else:
                                # 将 UTC 时间转换为本地时间后再移除时区信息
                                if parsed_dt.tzinfo is not None:
                                    # 转换为配置的时区
                                    parsed_dt = parsed_dt.astimezone(APP_TIMEZONE)
                                    # 移除时区信息
                                    parsed_dt = parsed_dt.replace(tzinfo=None)
                                converted[key] = parsed_dt
                        else:
                            # 无法解析，保持原值
                            converted[key] = value
                        continue
                    except Exception:
                        converted[key] = value
                        continue
                
                # 5. 尝试解析 JSON 类型（json, jsonb）
                if field_type in ('json', 'jsonb'):
                    try:
                        converted[key] = json.loads(value)
                        continue
                    except (json.JSONDecodeError, TypeError):
                        # 如果解析失败，保持原字符串
                        pass
                
                # 6. 数组类型（PostgreSQL array）
                if field_type.endswith('[]') or field_type in ('array', 'text[]', 'varchar[]', 'integer[]'):
                    try:
                        parsed = json.loads(value)
                        if isinstance(parsed, list):
                            converted[key] = parsed
                            continue
                    except (json.JSONDecodeError, TypeError):
                        # 尝试按逗号分隔
                        if ',' in value:
                            converted[key] = [item.strip() for item in value.split(',')]
                            continue
                
                # 默认保持原值
                converted[key] = value
            else:
                converted[key] = value
        
        return converted
    
    def _get_field_types(self) -> Dict[str, str]:
        """获取字段名到类型的映射"""
        field_types = {}
        
        # 从 form_config 中获取字段类型
        form_config = self._form_config
        table_configs = form_config.get('tableConfigs', [])
        
        for table_config in table_configs:
            fields = table_config.get('fields', [])
            for field in fields:
                field_name = field.get('name', '')
                field_type = field.get('type', '')
                if field_name and field_type:
                    field_types[field_name] = field_type
        
        return field_types

    async def create(self, db: AsyncSession, data: Dict[str, Any]) -> Dict[str, Any]:
        """新增数据（含子表，事务）"""
        main_data = data.get("main") or {}
        sub_tables_data = data.get("sub_tables") or {}

        # 过滤主表字段
        allowed_main_fields = self._get_allowed_fields("main")
        filtered_main = self._filter_fields(main_data, allowed_main_fields)
        
        # 转换数据类型
        filtered_main = self._convert_data_types(filtered_main)

        # 移除 id 字段，生成新的 UUID
        filtered_main.pop("id", None)
        generated_id = str(uuid.uuid4())
        filtered_main["id"] = generated_id
        
        # 填充系统字段（创建时间、创建人、部门等）
        filtered_main = self._fill_system_fields_for_create(filtered_main)

        if not filtered_main:
            raise FormDataException("主表数据不能为空")

        # 唯一性校验（新增时不需要排除ID）
        await self._validate_unique_fields(db, filtered_main)

        # 1. 插入主表
        table = self.form_meta.main_table
        schema = await self._resolve_schema(db, self.form_meta.main_table_schema) or None
        database = self.form_meta.main_table_database or None

        sql, params = self.sql_builder.build_insert(
            table=table,
            data=filtered_main,
            schema=schema,
            database=database,
            return_id=False
        )

        await self._execute_command(db, sql, params)
        main_pk = generated_id

        # 2. 插入子表
        for sub_table in self.sub_tables:
            sub_data_list = sub_tables_data.get(sub_table.table_name, [])
            if not sub_data_list:
                continue

            allowed_sub_fields = self._get_allowed_fields("sub", sub_table.table_name)

            for sub_item in sub_data_list:
                filtered_sub = self._filter_fields(sub_item, allowed_sub_fields)
                filtered_sub = self._convert_data_types(filtered_sub)
                filtered_sub.pop("id", None)
                filtered_sub["id"] = str(uuid.uuid4())
                filtered_sub[sub_table.foreign_key] = main_pk
                
                # 填充子表系统字段
                filtered_sub = self._fill_system_fields_for_create(filtered_sub)

                if filtered_sub:
                    sub_sql, sub_params = self.sql_builder.build_insert(
                        table=sub_table.table_name,
                        data=filtered_sub,
                        schema=sub_table.table_schema or None,
                        database=sub_table.table_database or None,
                        return_id=False
                    )
                    await self._execute_command(db, sub_sql, sub_params)

        await db.commit()

        logger.info(f"表单数据创建成功: form={self.form_meta.code}, pk={main_pk}")

        return await self.get(db, main_pk)

    async def update(self, db: AsyncSession, pk: Any, data: Dict[str, Any]) -> Dict[str, Any]:
        """更新数据（含子表，事务）"""
        main_data = data.get("main") or {}
        sub_tables_data = data.get("sub_tables") or {}

        # 验证数据存在
        existing = await self.get(db, pk)
        if not existing:
            raise FormDataException(f"数据不存在: {pk}")

        # 1. 更新主表
        if main_data:
            allowed_main_fields = self._get_allowed_fields("main")
            filtered_main = self._filter_fields(main_data, allowed_main_fields)
            filtered_main = self._convert_data_types(filtered_main)
            filtered_main.pop("id", None)
            
            # 填充系统字段（更新时间、修改人）
            filtered_main = self._fill_system_fields_for_update(filtered_main)

            # 唯一性校验（编辑时排除自身）
            await self._validate_unique_fields(db, filtered_main, exclude_id=str(pk))

            if filtered_main:
                table = self.form_meta.main_table
                schema = await self._resolve_schema(db, self.form_meta.main_table_schema) or None
                database = self.form_meta.main_table_database or None

                sql, params = self.sql_builder.build_update(
                    table=table,
                    data=filtered_main,
                    pk_field="id",
                    pk_value=pk,
                    schema=schema,
                    database=database
                )
                await self._execute_command(db, sql, params)

        # 2. 处理子表（差异更新）
        for sub_table in self.sub_tables:
            if sub_table.table_name not in sub_tables_data:
                continue

            new_sub_data = sub_tables_data[sub_table.table_name]
            await self._handle_sub_table_update(db, sub_table, pk, new_sub_data)

        await db.commit()

        logger.info(f"表单数据更新成功: form={self.form_meta.code}, pk={pk}")

        return await self.get(db, pk)

    async def _handle_sub_table_update(
            self,
            db: AsyncSession,
            sub_table: FormSubTable,
            main_pk: Any,
            new_data: List[Dict[str, Any]]
    ):
        """处理子表更新（差异对比：新增/更新/删除）"""
        table_name = sub_table.table_name
        schema = sub_table.table_schema or None
        database = sub_table.table_database or None
        foreign_key = sub_table.foreign_key

        # 获取现有数据
        existing = await self._query_sub_table_data(db, sub_table, main_pk)
        existing_map = {item["id"]: item for item in existing if "id" in item}

        allowed_fields = self._get_allowed_fields("sub", table_name)

        # 分类处理
        new_ids = set()
        to_insert = []
        to_update = []

        for item in new_data:
            item_id = item.get("id")
            if item_id and item_id in existing_map:
                new_ids.add(item_id)
                to_update.append(item)
            elif not item_id:
                to_insert.append(item)
            else:
                to_insert.append(item)

        # 找出需要删除的
        to_delete = [eid for eid in existing_map.keys() if eid not in new_ids]

        # 执行删除
        for del_id in to_delete:
            sql, params = self.sql_builder.build_delete(
                table=table_name,
                pk_field="id",
                pk_value=del_id,
                schema=schema,
                database=database
            )
            await self._execute_command(db, sql, params)

        # 执行更新
        for item in to_update:
            filtered = self._filter_fields(item, allowed_fields)
            filtered = self._convert_data_types(filtered)
            item_id = filtered.pop("id", None)
            
            # 填充子表系统字段（更新）
            filtered = self._fill_system_fields_for_update(filtered)
            
            if filtered and item_id:
                sql, params = self.sql_builder.build_update(
                    table=table_name,
                    data=filtered,
                    pk_field="id",
                    pk_value=item_id,
                    schema=schema,
                    database=database
                )
                await self._execute_command(db, sql, params)

        # 执行新增
        for item in to_insert:
            filtered = self._filter_fields(item, allowed_fields)
            filtered = self._convert_data_types(filtered)
            if "id" in filtered:
                del filtered["id"]
            filtered["id"] = str(uuid.uuid4())
            if foreign_key in filtered:
                del filtered[foreign_key]
            filtered[foreign_key] = main_pk
            
            # 填充子表系统字段（新增）
            filtered = self._fill_system_fields_for_create(filtered)

            if filtered and len(filtered) > 2:
                sql, params = self.sql_builder.build_insert(
                    table=table_name,
                    data=filtered,
                    schema=schema,
                    database=database,
                    return_id=False
                )
                await self._execute_command(db, sql, params)

    async def delete(self, db: AsyncSession, pk: Any) -> bool:
        """删除数据（含子表，事务）"""
        # 验证数据存在
        existing = await self.get(db, pk)
        if not existing:
            raise FormDataException(f"数据不存在: {pk}")

        # 1. 删除子表数据
        for sub_table in self.sub_tables:
            sql, params = self.sql_builder.build_delete_by_foreign_key(
                table=sub_table.table_name,
                fk_field=sub_table.foreign_key,
                fk_value=pk,
                schema=sub_table.table_schema or None,
                database=sub_table.table_database or None
            )
            await self._execute_command(db, sql, params)

        # 2. 删除主表数据
        table = self.form_meta.main_table
        schema = await self._resolve_schema(db, self.form_meta.main_table_schema) or None
        database = self.form_meta.main_table_database or None

        sql, params = self.sql_builder.build_delete(
            table=table,
            pk_field="id",
            pk_value=pk,
            schema=schema,
            database=database
        )
        affected = await self._execute_command(db, sql, params)

        await db.commit()

        logger.info(f"表单数据删除成功: form={self.form_meta.code}, pk={pk}")

        return affected > 0

    async def batch_delete(self, db: AsyncSession, pks: List[Any]) -> int:
        """
        批量删除（优化版：使用 IN 子句批量删除，单次事务提交）
        
        Args:
            db: 数据库会话
            pks: 主键列表
        
        Returns:
            成功删除的数量
        """
        if not pks:
            return 0
        
        # 1. 批量删除所有子表数据（使用 IN 子句）
        for sub_table in self.sub_tables:
            schema = await self._resolve_schema(db, sub_table.table_schema) if sub_table.table_schema else None
            database = sub_table.table_database or None
            
            # 构建批量删除 SQL: DELETE FROM table WHERE fk_field IN (:pk0, :pk1, ...)
            full_table = self.sql_builder.build_table_name(sub_table.table_name, schema, database)
            fk_field = self.sql_builder.quote_identifier(sub_table.foreign_key)
            
            placeholders = ", ".join(f":pk{i}" for i in range(len(pks)))
            sql = f"DELETE FROM {full_table} WHERE {fk_field} IN ({placeholders})"
            params = {f"pk{i}": pk for i, pk in enumerate(pks)}
            
            await self._execute_command(db, sql, params)
        
        # 2. 批量删除主表数据（使用 IN 子句）
        table = self.form_meta.main_table
        schema = await self._resolve_schema(db, self.form_meta.main_table_schema) or None
        database = self.form_meta.main_table_database or None
        
        full_table = self.sql_builder.build_table_name(table, schema, database)
        id_field = self.sql_builder.quote_identifier("id")
        
        placeholders = ", ".join(f":pk{i}" for i in range(len(pks)))
        sql = f"DELETE FROM {full_table} WHERE {id_field} IN ({placeholders})"
        params = {f"pk{i}": pk for i, pk in enumerate(pks)}
        
        affected = await self._execute_command(db, sql, params)
        
        # 3. 单次提交
        await db.commit()
        
        logger.info(f"批量删除成功: form={self.form_meta.code}, count={affected}")
        
        return affected

    # ============ 工具方法 ============

    def _serialize_row(self, row: Dict[str, Any]) -> Dict[str, Any]:
        """序列化行数据（处理特殊类型）"""
        result = {}
        for key, value in row.items():
            if isinstance(value, datetime):
                result[key] = value.strftime("%Y-%m-%d %H:%M:%S")
            elif isinstance(value, date):
                result[key] = value.strftime("%Y-%m-%d")
            elif isinstance(value, Decimal):
                result[key] = float(value)
            elif isinstance(value, uuid.UUID):
                result[key] = str(value)
            elif isinstance(value, bytes):
                result[key] = value.decode("utf-8", errors="ignore")
            elif isinstance(value, str):
                result[key] = value.strip()
            else:
                result[key] = value
        return result

    # ============ 导入导出 ============

    async def export_to_excel_streaming(
            self,
            db: AsyncSession,
            selected_fields: List[str] = None,
            include_sub_tables: bool = False,
            batch_size: int = 1000,
            data_scope: Dict[str, Any] = None,
            filters: Dict[str, Any] = None,
            sort_list: List[Dict[str, str]] = None,
            search: str = None,
            search_fields: List[str] = None,
            on_progress: Any = None
    ) -> BytesIO:
        """
        流式导出数据到 Excel（分批查询，避免内存溢出）
        
        Args:
            db: 数据库会话
            selected_fields: 选中的字段列表
            include_sub_tables: 是否导出子表
            batch_size: 每批查询的数据量
            data_scope: 数据权限过滤配置
            filters: 过滤条件（与列表查询一致）
            sort_list: 排序列表
            search: 搜索关键词
            search_fields: 搜索字段列表
            on_progress: 进度回调函数 async fn(processed, total)，可选
        
        Returns:
            Excel 文件的 BytesIO 对象
        """
        wb = Workbook()
        ws = wb.active
        ws.title = "主表数据"

        # 从列表配置中获取列定义
        list_config = self._list_config
        columns = list_config.get("columns", [])

        # 如果指定了选择的字段，只导出这些字段
        if selected_fields and columns:
            columns = [col for col in columns if col.get("field") in selected_fields]
            # 确保 selected_fields 中的字段都被包含
            existing_fields = {col.get("field") for col in columns}
            for field in selected_fields:
                if field not in existing_fields:
                    columns.append({"field": field, "label": field.upper()})

        if columns:
            headers = [col.get("field") for col in columns]
            column_labels = [col.get("label", col.get("field")) for col in columns]
            display_fields = {
                col.get("field"): col.get("displayField")
                for col in columns
                if col.get("displayField")
            }
        else:
            # 如果没有列配置，先查询一条数据获取字段
            first_result = await self.list(db=db, page=1, page_size=1, data_scope=data_scope)
            if first_result["items"]:
                headers = list(first_result["items"][0].keys())
            else:
                headers = []
            column_labels = headers
            display_fields = {}

        # 从 relation_fields 补充缺少 displayField 的关联字段映射
        relation_fields = self._get_relation_fields()
        for field_name, config in relation_fields.items():
            if field_name not in display_fields and config.get("display_field"):
                display_fields[field_name] = config["display_field"]

        # 先查询总数（用于进度计算）
        total_count = 0
        if on_progress:
            count_result = await self.list(
                db=db, page=1, page_size=1, data_scope=data_scope,
                filters=filters, sort_list=sort_list,
                search=search, search_fields=search_fields
            )
            total_count = min(count_result.get("total", 0), MAX_IMPORT_EXPORT_ROWS)
            await on_progress(0, total_count, "querying")

        # 定义样式
        header_font = Font(name='微软雅黑', size=11, bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center')
        data_alignment = Alignment(horizontal='left', vertical='center')
        thin_border = Border(
            left=Side(style='thin', color='D0D0D0'),
            right=Side(style='thin', color='D0D0D0'),
            top=Side(style='thin', color='D0D0D0'),
            bottom=Side(style='thin', color='D0D0D0')
        )

        # 写入表头
        ws.append(column_labels)
        for col_idx, cell in enumerate(ws[1], 1):
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
            column_letter = cell.column_letter
            header_length = len(str(cell.value))
            ws.column_dimensions[column_letter].width = max(12, header_length + 4)

        # 子表 sheet 初始化（延迟到第一批有数据时创建表头）
        sub_sheets: Dict[str, Any] = {}
        sub_headers_map: Dict[str, List[str]] = {}

        # 分批查询并写入数据（主表 + 子表同步处理）
        page = 1
        total_rows = 0

        while True:
            result = await self.list(
                db=db, page=page, page_size=batch_size, data_scope=data_scope,
                filters=filters, sort_list=sort_list,
                search=search, search_fields=search_fields
            )
            items = result["items"]

            if not items:
                break

            if page == 1 and total_count == 0:
                total_count = min(result.get("total", 0), MAX_IMPORT_EXPORT_ROWS)

            batch_ids = []
            for item in items:
                row = []
                for header in headers:
                    display_field = display_fields.get(header)
                    if display_field and display_field in item:
                        value = item.get(display_field, "")
                    else:
                        value = item.get(header, "")
                    if isinstance(value, (list, dict)):
                        value = str(value)
                    row.append(value)
                ws.append(row)
                total_rows += 1
                if include_sub_tables and item.get("id"):
                    batch_ids.append(item.get("id"))

            # 当前批次的子表数据立即查询并写入
            if include_sub_tables and self.sub_tables and batch_ids:
                for sub_table in self.sub_tables:
                    sub_key = sub_table.alias or sub_table.table_name
                    sql, params = self.sql_builder.build_select(
                        table=sub_table.table_name,
                        schema=sub_table.table_schema or None,
                        database=sub_table.table_database or None,
                        where={sub_table.foreign_key: {"type": "in", "value": batch_ids}}
                    )
                    rows = await self._execute_query(db, sql, params)
                    if not rows:
                        continue

                    sub_items = [self._serialize_row(r) for r in rows]

                    if sub_key not in sub_sheets:
                        ws_sub = wb.create_sheet(title=sub_key[:31])
                        sub_hdrs = ["主表ID"] + [k for k in sub_items[0].keys()]
                        ws_sub.append(sub_hdrs)
                        for col_idx, cell in enumerate(ws_sub[1], 1):
                            cell.font = header_font
                            cell.fill = header_fill
                            cell.alignment = header_alignment
                            cell.border = thin_border
                            ws_sub.column_dimensions[cell.column_letter].width = max(12, len(str(cell.value)) + 4)
                        sub_sheets[sub_key] = ws_sub
                        sub_headers_map[sub_key] = sub_hdrs
                    else:
                        ws_sub = sub_sheets[sub_key]
                        sub_hdrs = sub_headers_map[sub_key]

                    for sub_item in sub_items:
                        row = [sub_item.get(sub_table.foreign_key, "")]
                        for h in sub_hdrs[1:]:
                            value = sub_item.get(h, "")
                            if isinstance(value, (list, dict)):
                                value = str(value)
                            row.append(value)
                        ws_sub.append(row)

            if on_progress:
                await on_progress(total_rows, total_count, "querying")

            if len(items) < batch_size:
                break
            page += 1
            if total_rows >= MAX_IMPORT_EXPORT_ROWS:
                logger.warning(f"导出数据量达到上限 {MAX_IMPORT_EXPORT_ROWS} 条（服务器内存 {SERVER_MEMORY_GB}GB），停止导出")
                break

        if on_progress:
            await on_progress(total_rows, total_count, "generating")

        ws.freeze_panes = 'A2'
        for ws_sub in sub_sheets.values():
            ws_sub.freeze_panes = 'A2'

        # 保存到内存
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        logger.info(f"流式导出完成: 共 {total_rows} 条数据")
        return output

    async def export_to_excel(self, items: List[Dict[str, Any]], selected_fields: List[str] = None, sub_tables_data: Dict[str, List[Dict[str, Any]]] = None) -> BytesIO:
        """导出数据到 Excel（带样式，支持字段选择和子表导出）"""
        wb = Workbook()
        ws = wb.active
        ws.title = "主表数据"

        if not items:
            output = BytesIO()
            wb.save(output)
            output.seek(0)
            return output

        # 从列表配置中获取列定义
        list_config = self._list_config
        columns = list_config.get("columns", [])

        # 如果没有列配置，使用第一行数据的所有字段
        if not columns:
            headers = list(items[0].keys())
            column_labels = headers
            display_fields = {}
        else:
            # 如果指定了选择的字段，只导出这些字段
            if selected_fields:
                columns = [col for col in columns if col.get("field") in selected_fields]
                
                # 确保 selected_fields 中的字段都被包含（即使不在 columns 配置中）
                existing_fields = {col.get("field") for col in columns}
                for field in selected_fields:
                    if field not in existing_fields:
                        # 添加缺失的字段（如 id）
                        columns.append({"field": field, "label": field.upper()})
            
            headers = [col.get("field") for col in columns]
            column_labels = [col.get("label", col.get("field")) for col in columns]
            # 构建字段到 displayField 的映射
            display_fields = {
                col.get("field"): col.get("displayField")
                for col in columns
                if col.get("displayField")
            }

        # 定义样式
        # 表头样式：深蓝色背景，白色粗体文字
        header_font = Font(name='微软雅黑', size=11, bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center')
        
        # 数据单元格样式：居中对齐
        data_alignment = Alignment(horizontal='left', vertical='center')
        
        # 边框样式
        thin_border = Border(
            left=Side(style='thin', color='D0D0D0'),
            right=Side(style='thin', color='D0D0D0'),
            top=Side(style='thin', color='D0D0D0'),
            bottom=Side(style='thin', color='D0D0D0')
        )

        # 写入表头
        ws.append(column_labels)
        
        # 设置表头样式
        for col_idx, cell in enumerate(ws[1], 1):
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
            # 设置列宽（根据表头长度自动调整）
            column_letter = cell.column_letter
            header_length = len(str(cell.value))
            ws.column_dimensions[column_letter].width = max(12, header_length + 4)

        # 写入数据
        for item in items:
            row = []
            for header in headers:
                # 优先使用 displayField 的值（如果配置了）
                display_field = display_fields.get(header)
                if display_field and display_field in item:
                    value = item.get(display_field, "")
                else:
                    value = item.get(header, "")
                
                # 处理特殊类型
                if isinstance(value, (list, dict)):
                    value = str(value)
                row.append(value)
            ws.append(row)

        # 设置数据行样式
        for row_idx in range(2, ws.max_row + 1):
            for col_idx in range(1, ws.max_column + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.alignment = data_alignment
                cell.border = thin_border
                # 交替行背景色（浅灰色）
                if row_idx % 2 == 0:
                    cell.fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')

        # 冻结首行（表头）
        ws.freeze_panes = 'A2'

        # 导出子表数据
        if sub_tables_data:
            for sub_table_name, sub_data_list in sub_tables_data.items():
                if not sub_data_list:
                    continue
                
                # 为每个子表创建新的工作表
                ws_sub = wb.create_sheet(title=sub_table_name[:31])  # Excel 工作表名称最多 31 字符
                
                # 子表表头（包含主表ID）
                sub_headers = ["主表ID"] + [k for k in sub_data_list[0].keys() if k != "_main_id"]
                ws_sub.append(sub_headers)
                
                # 设置子表表头样式
                for col_idx, cell in enumerate(ws_sub[1], 1):
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = header_alignment
                    cell.border = thin_border
                    column_letter = cell.column_letter
                    ws_sub.column_dimensions[column_letter].width = max(12, len(str(cell.value)) + 4)
                
                # 写入子表数据
                for sub_item in sub_data_list:
                    row = [sub_item.get("_main_id", "")]
                    for header in sub_headers[1:]:
                        value = sub_item.get(header, "")
                        # 处理特殊类型
                        if isinstance(value, (list, dict)):
                            value = str(value)
                        row.append(value)
                    ws_sub.append(row)
                
                # 设置子表数据行样式
                for row_idx in range(2, ws_sub.max_row + 1):
                    for col_idx in range(1, ws_sub.max_column + 1):
                        cell = ws_sub.cell(row=row_idx, column=col_idx)
                        cell.alignment = data_alignment
                        cell.border = thin_border
                        if row_idx % 2 == 0:
                            cell.fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
                
                ws_sub.freeze_panes = 'A2'

        # 保存到内存
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output

    async def get_import_template(self) -> BytesIO:
        """生成导入模板（带样式）"""
        wb = Workbook()
        ws = wb.active
        ws.title = "导入模板"

        # 从列表配置中获取列定义
        list_config = self._list_config
        columns = list_config.get("columns", [])

        # 定义样式
        header_font = Font(name='微软雅黑', size=11, bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center')
        thin_border = Border(
            left=Side(style='thin', color='D0D0D0'),
            right=Side(style='thin', color='D0D0D0'),
            top=Side(style='thin', color='D0D0D0'),
            bottom=Side(style='thin', color='D0D0D0')
        )

        if columns:
            # 使用列配置生成表头（排除系统字段和关联字段的 displayField）
            headers = []
            for col in columns:
                field = col.get("field")
                # 排除系统字段
                if field not in ["id", "created_at", "updated_at", "sys_create_datetime", "sys_update_datetime"]:
                    # 排除 displayField（如 dept_name），只保留原始字段（如 dept_id）
                    display_field = col.get("displayField")
                    if not display_field or field != display_field:
                        headers.append(col.get("label", field))
            
            ws.append(headers)
            
            # 设置表头样式
            for col_idx, cell in enumerate(ws[1], 1):
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = thin_border
                # 设置列宽
                column_letter = cell.column_letter
                header_length = len(str(cell.value))
                ws.column_dimensions[column_letter].width = max(15, header_length + 4)
            
            # 添加示例数据行（带浅黄色背景提示这是示例）
            example_row = ["示例数据，请删除此行后填写实际数据"] + [""] * (len(headers) - 1)
            ws.append(example_row)
            for col_idx in range(1, len(headers) + 1):
                cell = ws.cell(row=2, column=col_idx)
                cell.fill = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
                cell.alignment = Alignment(horizontal='left', vertical='center')
                cell.border = thin_border
                cell.font = Font(name='微软雅黑', size=10, italic=True, color='808080')
        else:
            # 如果没有列配置，添加提示
            ws.append(["请配置列表字段后重新下载模板"])
            ws['A1'].font = Font(name='微软雅黑', size=12, bold=True, color='FF0000')
            ws.column_dimensions['A'].width = 40

        # 冻结首行
        ws.freeze_panes = 'A2'

        # 保存到内存
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output

    async def import_from_excel(
            self, 
            db: AsyncSession, 
            file_content: bytes,
            batch_size: int = 1000,
            mode: str = "append",
            validate_only: bool = False,
            data_handling: str = "insert_only",
            match_field: str = None,
            on_progress=None
    ) -> tuple[int, int, List[Dict[str, Any]]]:
        """
        从 Excel 批量导入数据
        
        Args:
            db: 数据库会话
            file_content: Excel 文件内容
            batch_size: 批量插入的批次大小，默认 1000
            mode: 导入模式，"append"（追加）或 "overwrite"（覆盖，先清空表再导入）
            validate_only: 是否仅验证数据，不执行实际导入
            data_handling: 数据处理方式 insert_only / update_only / upsert
            match_field: 更新模式下用于匹配的字段名
            on_progress: 可选的异步回调 (processed, total, stage, success, fail)
        
        Returns:
            (成功数量, 失败数量, 错误详情列表)
        """
        from utils.context import get_current_user_info_from_context
        from sqlalchemy import text as sa_text
        
        file_io = BytesIO(file_content)
        del file_content
        wb = load_workbook(file_io, read_only=True)
        ws = wb.active

        # read_only 模式下先读表头
        row_iter = ws.iter_rows(values_only=True)
        headers = list(next(row_iter, []))

        # read_only 模式 ws.max_row 可能不准，用预估值（后续会动态修正）
        total_excel_rows = (ws.max_row - 1) if ws.max_row and ws.max_row > 1 else 0

        # 从列表配置中获取字段映射
        list_config = self._list_config
        columns = list_config.get("columns", [])
        label_to_field = {col.get("label"): col.get("field") for col in columns if col.get("label") and col.get("field")}

        # 获取允许的字段
        allowed_fields = self._get_allowed_fields("main")
        
        # 系统字段（需要排除）
        system_fields = {"id", "created_at", "updated_at", "sys_create_datetime", "sys_update_datetime", 
                         "sys_creator_id", "sys_modifier_id", "sys_dept_id", "is_deleted", "sort"}

        success_count = 0
        fail_count = 0
        error_details = []
        
        # 获取用户信息（用于填充系统字段）
        user_info = get_current_user_info_from_context()
        
        if on_progress:
            await on_progress(0, total_excel_rows, "parsing", 0, 0)

        # 根据预估总行数动态计算进度推送间隔（约推送 50~100 次，最少 50 行，最多 5000 行）
        progress_interval = max(50, min(5000, total_excel_rows // 100)) if total_excel_rows > 0 else 100

        # 流式逐行解析（read_only 模式下 iter_rows 是惰性的）
        all_rows = []
        parsed_count = 0
        for row in row_iter:
            row_num = parsed_count + 2
            try:
                # 跳过空行
                if all(cell is None or (isinstance(cell, str) and cell.strip() == "") for cell in row):
                    parsed_count += 1
                    continue
                
                # 构建数据字典
                data = {}
                for idx, value in enumerate(row):
                    if idx < len(headers) and headers[idx]:
                        field = label_to_field.get(headers[idx], headers[idx])
                        if field and field not in system_fields:
                            if field in allowed_fields:
                                data[field] = value

                if data:
                    data = self._convert_data_types(data)
                    data = self._filter_fields(data, allowed_fields)
                    
                    is_update_mode = mode == "append" and data_handling in ("update_only", "upsert")
                    if not is_update_mode:
                        data["id"] = str(uuid.uuid4())
                        data = self._fill_system_fields_for_create(data)
                    
                    data = self._normalize_data_for_insert(data)
                    
                    validation_error = self._validate_data_against_schema(row_num, data)
                    if validation_error:
                        fail_count += 1
                        error_details.append({"row": row_num, "error": validation_error})
                        parsed_count += 1
                        continue
                    
                    all_rows.append((row_num, data))
                    if len(all_rows) >= MAX_IMPORT_EXPORT_ROWS:
                        logger.warning(f"导入数据量达到上限 {MAX_IMPORT_EXPORT_ROWS} 条（服务器内存 {SERVER_MEMORY_GB}GB），截断后续数据")
                        parsed_count += 1
                        break
            except Exception as e:
                logger.error(f"解析第 {row_num} 行数据失败: {e}")
                fail_count += 1
                error_details.append({"row": row_num, "error": str(e)})
            
            parsed_count += 1
            if on_progress and parsed_count % progress_interval == 0:
                if parsed_count > total_excel_rows:
                    total_excel_rows = parsed_count
                    progress_interval = max(50, min(5000, total_excel_rows // 100))
                await on_progress(parsed_count, total_excel_rows, "parsing", 0, fail_count)
                await asyncio.sleep(0)
        
        if parsed_count != total_excel_rows:
            total_excel_rows = parsed_count
        if on_progress:
            await on_progress(parsed_count, total_excel_rows, "parsing", 0, fail_count)
        
        wb.close()
        file_io.close()
        del wb, ws, row_iter, file_io

        if not all_rows:
            return success_count, fail_count, error_details
        
        # 预先解析 schema（只解析一次）
        table = self.form_meta.main_table
        schema = await self._resolve_schema(db, self.form_meta.main_table_schema) or None
        database = self.form_meta.main_table_database or None
        
        # **按导入模式执行不同的验证规则**
        if on_progress:
            await on_progress(0, len(all_rows), "validating", 0, fail_count)

        is_update_mode = mode == "append" and data_handling in ("update_only", "upsert")
        unique_errors: Dict[int, str] = {}
        existing_set: Set[str] = set()

        if mode == "overwrite":
            unique_errors = await self._batch_check_unique_fields_internal_only(all_rows)

        elif data_handling == "insert_only":
            unique_errors = await self._batch_check_unique_fields(db, all_rows)

        elif data_handling in ("update_only", "upsert") and match_field:
            match_label = self._get_field_labels([match_field]).get(match_field, match_field)

            for row_num, data in all_rows:
                val = data.get(match_field)
                if val is None or val == '':
                    unique_errors[row_num] = f"匹配字段 '{match_label}' 不能为空"

            match_val_to_rows: Dict[str, list] = {}
            for row_num, data in all_rows:
                val = data.get(match_field)
                if val is not None and val != '':
                    match_val_to_rows.setdefault(str(val), []).append(row_num)
            for val, rows in match_val_to_rows.items():
                if len(rows) > 1:
                    for r in rows[1:]:
                        if r not in unique_errors:
                            unique_errors[r] = f"匹配字段 '{match_label}' 值 '{val}' 在导入数据中重复（第 {rows[0]} 行已存在）"

            if on_progress:
                await on_progress(len(all_rows) // 3, len(all_rows), "validating", 0, fail_count)

            match_values = [str(data.get(match_field)) for _, data in all_rows
                            if data.get(match_field) not in (None, '')]
            existing_set = await self._batch_query_existing(db, match_field, match_values)

            if on_progress:
                await on_progress(len(all_rows) * 2 // 3, len(all_rows), "validating", 0, fail_count)

            if data_handling == "update_only":
                for row_num, data in all_rows:
                    val = data.get(match_field)
                    if val is not None and val != '' and str(val) not in existing_set and row_num not in unique_errors:
                        unique_errors[row_num] = f"未找到匹配记录（{match_label}={val}）"
            else:
                new_rows = [(rn, d) for rn, d in all_rows
                            if d.get(match_field) is not None and d.get(match_field) != ''
                            and str(d.get(match_field)) not in existing_set
                            and rn not in unique_errors]
                if new_rows:
                    insert_unique_errors = await self._batch_check_unique_fields(db, new_rows)
                    unique_errors.update(insert_unique_errors)

        if on_progress:
            await on_progress(len(all_rows), len(all_rows), "validating", 0, fail_count)
        
        if unique_errors:
            filtered_rows = []
            for row_num, data in all_rows:
                if row_num in unique_errors:
                    fail_count += 1
                    error_details.append({"row": row_num, "error": unique_errors[row_num]})
                else:
                    filtered_rows.append((row_num, data))
            all_rows = filtered_rows
        
        if not all_rows:
            logger.info(f"Excel 导入完成: 成功 {success_count} 条, 失败 {fail_count} 条（所有数据验证失败）")
            return success_count, fail_count, error_details[:100]
        
        # **仅验证模式**：返回验证结果，不执行实际导入
        if validate_only:
            success_count = len(all_rows)
            # 计算各模式下的操作预估
            validate_meta: Dict[str, Any] = {"_meta": True}
            if mode == "overwrite":
                validate_meta["will_insert"] = success_count
                validate_meta["will_update"] = 0
                validate_meta["action"] = "overwrite"
            elif data_handling == "insert_only":
                validate_meta["will_insert"] = success_count
                validate_meta["will_update"] = 0
                validate_meta["action"] = "insert_only"
            elif data_handling == "update_only":
                validate_meta["will_insert"] = 0
                validate_meta["will_update"] = success_count
                validate_meta["action"] = "update_only"
            elif data_handling == "upsert" and match_field:
                will_update = sum(
                    1 for _, d in all_rows
                    if d.get(match_field) is not None and d.get(match_field) != ''
                    and str(d.get(match_field)) in existing_set
                )
                will_insert = success_count - will_update
                validate_meta["will_insert"] = will_insert
                validate_meta["will_update"] = will_update
                validate_meta["action"] = "upsert"
            else:
                validate_meta["will_insert"] = success_count
                validate_meta["will_update"] = 0
                validate_meta["action"] = data_handling

            truncated_errors = error_details[:100]
            truncated_errors.append(validate_meta)
            logger.info(f"Excel 数据验证完成: 通过 {success_count} 条, 失败 {fail_count} 条, "
                        f"预计新增 {validate_meta.get('will_insert', 0)} 条, "
                        f"预计更新 {validate_meta.get('will_update', 0)} 条")
            return success_count, fail_count, truncated_errors
        
        # **覆盖模式**：先清空表数据
        if mode == "overwrite":
            await self._truncate_table(db, table, schema, database)
            logger.info(f"覆盖模式：已清空表 {table}")
        
        total_valid = len(all_rows)
        imported_count = 0

        # 根据字段数动态调整 batch_size，避免超出 PostgreSQL 32767 参数限制
        if all_rows:
            num_fields = len(all_rows[0][1])
            if num_fields > 0:
                max_safe_batch = 32000 // num_fields
                batch_size = max(1, min(batch_size, max_safe_batch))

        if on_progress:
            await on_progress(0, total_valid, "importing", success_count, fail_count)

        # **更新/upsert 模式**
        if is_update_mode and match_field:
            success_count, fail_count, error_details = await self._import_with_update(
                db, table, schema, database, all_rows, match_field,
                data_handling, success_count, fail_count, error_details,
                on_progress=on_progress
            )
            return success_count, fail_count, error_details[:100]
        
        # **纯插入模式**：分批处理
        for batch_start in range(0, len(all_rows), batch_size):
            batch = all_rows[batch_start:batch_start + batch_size]
            batch_data_list = [item[1] for item in batch]
            batch_row_nums = [item[0] for item in batch]
            
            sp_batch = f"sp_batch_{uuid.uuid4().hex[:8]}"
            try:
                await db.execute(sa_text(f"SAVEPOINT {sp_batch}"))
                await self._batch_insert_raw(db, table, schema, database, batch_data_list)
                await db.execute(sa_text(f"RELEASE SAVEPOINT {sp_batch}"))
                success_count += len(batch_data_list)
            except Exception as batch_err:
                logger.warning(f"批量插入失败（第 {batch_row_nums[0]}-{batch_row_nums[-1]} 行），降级逐条插入: {batch_err}")
                await db.execute(sa_text(f"ROLLBACK TO SAVEPOINT {sp_batch}"))
                
                for row_num, data in batch:
                    sp_single = f"sp_{uuid.uuid4().hex[:8]}"
                    try:
                        await db.execute(sa_text(f"SAVEPOINT {sp_single}"))
                        sql, params = self.sql_builder.build_insert(
                            table=table, data=data, schema=schema,
                            database=database, return_id=False
                        )
                        await self._execute_command(db, sql, params)
                        await db.execute(sa_text(f"RELEASE SAVEPOINT {sp_single}"))
                        success_count += 1
                    except Exception as single_err:
                        await db.execute(sa_text(f"ROLLBACK TO SAVEPOINT {sp_single}"))
                        fail_count += 1
                        err_msg = str(single_err)
                        if "DETAIL:" in err_msg:
                            err_msg = err_msg.split("DETAIL:")[1].split("[SQL")[0].strip()
                        elif "duplicate key" in err_msg:
                            err_msg = err_msg.split(")")[0] + ")" if ")" in err_msg else err_msg[:200]
                        else:
                            err_msg = err_msg[:200]
                        error_details.append({"row": row_num, "error": err_msg})
            
            await db.commit()
            db.expire_all()
            imported_count += len(batch)
            if on_progress:
                await on_progress(imported_count, total_valid, "importing", success_count, fail_count)

        logger.info(f"Excel 导入完成: 成功 {success_count} 条, 失败 {fail_count} 条")
        return success_count, fail_count, error_details[:100]

    async def _import_with_update(
            self,
            db: AsyncSession,
            table: str,
            schema: Optional[str],
            database: Optional[str],
            all_rows: List[tuple],
            match_field: str,
            data_handling: str,
            success_count: int,
            fail_count: int,
            error_details: List[Dict[str, Any]],
            on_progress=None
    ) -> tuple[int, int, List[Dict[str, Any]]]:
        """
        导入数据（更新/upsert 模式）
        
        通过 match_field 匹配已有记录：
        - update_only: 匹配到则更新，未匹配到则跳过
        - upsert: 匹配到则更新，未匹配到则新增
        """
        full_table = self.sql_builder.build_table_name(table, schema, database)
        quoted_match = self.sql_builder.quote_identifier(match_field)
        total_valid = len(all_rows)
        
        # 1. 收集所有行的匹配字段值
        match_values = []
        for row_num, data in all_rows:
            val = data.get(match_field)
            if val is not None:
                match_values.append(val)
        
        # 2. 批量查询数据库中已存在的记录（match_field -> id 的映射）
        existing_map = {}
        if match_values:
            for i in range(0, len(match_values), 500):
                batch_vals = match_values[i:i + 500]
                placeholders = ", ".join([f":mv_{j}" for j in range(len(batch_vals))])
                query_sql = f"SELECT \"id\", {quoted_match} FROM {full_table} WHERE {quoted_match} IN ({placeholders})"
                params = {f"mv_{j}": v for j, v in enumerate(batch_vals)}
                try:
                    result = await self._execute_query(db, query_sql, params)
                    for row in result:
                        existing_map[row[match_field]] = row["id"]
                except Exception as e:
                    logger.error(f"查询已有数据失败: {e}")
        
        # 3. 分批处理每行数据
        imported_count = 0
        commit_batch_size = 500
        import_progress_interval = max(50, min(5000, total_valid // 100)) if total_valid > 0 else 100
        for row_num, data in all_rows:
            match_val = data.get(match_field)
            existing_id = existing_map.get(match_val) if match_val is not None else None
            
            sp = f"sp_{uuid.uuid4().hex[:8]}"
            try:
                await db.execute(text(f"SAVEPOINT {sp}"))
                
                if existing_id:
                    update_data = {k: v for k, v in data.items() if k not in ("id", "sys_create_datetime", "sys_creator_id", "sys_dept_id")}
                    update_data = self._fill_system_fields_for_update(update_data)
                    
                    sql, params = self.sql_builder.build_update(
                        table=table, data=update_data, pk_field="id", pk_value=existing_id,
                        schema=schema, database=database
                    )
                    await self._execute_command(db, sql, params)
                    await db.execute(text(f"RELEASE SAVEPOINT {sp}"))
                    success_count += 1
                elif data_handling == "upsert":
                    data["id"] = str(uuid.uuid4())
                    data = self._fill_system_fields_for_create(data)
                    sql, params = self.sql_builder.build_insert(
                        table=table, data=data, schema=schema,
                        database=database, return_id=False
                    )
                    await self._execute_command(db, sql, params)
                    await db.execute(text(f"RELEASE SAVEPOINT {sp}"))
                    success_count += 1
                else:
                    await db.execute(text(f"RELEASE SAVEPOINT {sp}"))
                    fail_count += 1
                    error_details.append({"row": row_num, "error": f"未找到匹配记录（{match_field}={match_val}）"})
            except Exception as e:
                await db.execute(text(f"ROLLBACK TO SAVEPOINT {sp}"))
                fail_count += 1
                err_msg = str(e)
                if "DETAIL:" in err_msg:
                    err_msg = err_msg.split("DETAIL:")[1].split("[SQL")[0].strip()
                else:
                    err_msg = err_msg[:200]
                error_details.append({"row": row_num, "error": err_msg})
            
            imported_count += 1
            if imported_count % commit_batch_size == 0:
                await db.commit()
                db.expire_all()
            if on_progress and imported_count % import_progress_interval == 0:
                await on_progress(imported_count, total_valid, "importing", success_count, fail_count)
        
        await db.commit()
        db.expire_all()
        if on_progress:
            await on_progress(imported_count, total_valid, "importing", success_count, fail_count)
        logger.info(f"Excel 导入完成（{data_handling}）: 成功 {success_count} 条, 失败 {fail_count} 条")
        return success_count, fail_count, error_details

    async def _truncate_table(
            self, db: AsyncSession, table: str, schema: Optional[str], database: Optional[str]
    ) -> None:
        """
        清空表数据（用于覆盖模式）
        
        Args:
            db: 数据库会话
            table: 表名
            schema: Schema 名
            database: 数据库名
        """
        from sqlalchemy import text as sa_text
        
        full_table = self.sql_builder.build_table_name(table, schema, database)
        
        # 先删除子表数据
        for sub_table in self.sub_tables:
            sub_schema = await self._resolve_schema(db, sub_table.table_schema) if sub_table.table_schema else None
            sub_database = sub_table.table_database or None
            sub_full_table = self.sql_builder.build_table_name(sub_table.table_name, sub_schema, sub_database)
            await db.execute(sa_text(f"DELETE FROM {sub_full_table}"))
        
        # 再删除主表数据
        await db.execute(sa_text(f"DELETE FROM {full_table}"))
        await db.commit()

    async def _batch_check_unique_fields_internal_only(
            self, data_list: List[tuple[int, Dict[str, Any]]]
    ) -> Dict[int, str]:
        """
        仅检查导入数据内部重复（用于覆盖模式，跳过数据库检查）
        
        Args:
            data_list: [(row_num, data), ...] 数据列表
        
        Returns:
            {row_num: error_message} 违反唯一性约束的行号和错误信息
        """
        unique_fields = self._get_unique_check_fields()
        if not unique_fields:
            return {}
        
        errors = {}
        
        for field_name in unique_fields:
            # 收集该字段的所有非空值及其行号
            value_to_rows = {}  # {value: [row_num1, row_num2, ...]}
            for row_num, data in data_list:
                value = data.get(field_name)
                if value is None or value == '':
                    continue
                value_str = str(value)
                if value_str not in value_to_rows:
                    value_to_rows[value_str] = []
                value_to_rows[value_str].append(row_num)
            
            # 检查内部重复
            for value, row_nums in value_to_rows.items():
                if len(row_nums) > 1:
                    field_label = self._get_field_labels([field_name]).get(field_name, field_name)
                    for row_num in row_nums[1:]:
                        errors[row_num] = f"字段 '{field_label}' 值 '{value}' 在导入数据中重复（第 {row_nums[0]} 行已存在）"
        
        return errors

    async def _batch_check_unique_fields(
            self, db: AsyncSession, data_list: List[tuple[int, Dict[str, Any]]]
    ) -> Dict[int, str]:
        """
        批量检查唯一性约束（高性能版）
        
        Args:
            db: 数据库会话
            data_list: [(row_num, data), ...] 数据列表
        
        Returns:
            {row_num: error_message} 违反唯一性约束的行号和错误信息
        """
        from sqlalchemy import text as sa_text
        
        unique_fields = self._get_unique_check_fields()
        if not unique_fields:
            return {}
        
        errors = {}
        table = self.form_meta.main_table
        schema = await self._resolve_schema(db, self.form_meta.main_table_schema) or None
        database = self.form_meta.main_table_database or None
        table_name = self.sql_builder.build_table_name(table, schema, database)
        
        # 对每个唯一字段进行批量检查
        for field_name in unique_fields:
            # 收集该字段的所有非空值及其行号
            value_to_rows = {}  # {value: [row_num1, row_num2, ...]}
            for row_num, data in data_list:
                value = data.get(field_name)
                if value is None or value == '':
                    continue
                value_str = str(value)
                if value_str not in value_to_rows:
                    value_to_rows[value_str] = []
                value_to_rows[value_str].append(row_num)
            
            if not value_to_rows:
                continue
            
            # 1. 检查导入数据内部重复
            for value, row_nums in value_to_rows.items():
                if len(row_nums) > 1:
                    # 内部重复：除了第一行，其他行都标记为错误
                    field_label = self._get_field_labels([field_name]).get(field_name, field_name)
                    for row_num in row_nums[1:]:
                        errors[row_num] = f"字段 '{field_label}' 值 '{value}' 在导入数据中重复（第 {row_nums[0]} 行已存在）"
            
            # 2. 分批检查数据库中是否已存在（避免超出 32767 参数限制）
            values = list(value_to_rows.keys())
            if not values:
                continue
            
            quoted_field = self.sql_builder.quote_identifier(field_name)
            has_is_deleted = await self._check_column_exists(db, table, schema, database, "is_deleted")
            existing_values: set = set()

            for batch_start in range(0, len(values), 500):
                batch_vals = values[batch_start:batch_start + 500]
                placeholders = ", ".join(f":v{i}" for i in range(len(batch_vals)))
                sql = f"SELECT {quoted_field} FROM {table_name} WHERE {quoted_field} IN ({placeholders})"
                params = {f"v{i}": v for i, v in enumerate(batch_vals)}

                if has_is_deleted:
                    quoted_is_deleted = self.sql_builder.quote_identifier("is_deleted")
                    sql += f" AND {quoted_is_deleted} = :is_deleted"
                    params["is_deleted"] = False

                result = await db.execute(sa_text(sql), params)
                existing_values.update(str(row[0]) for row in result.fetchall())
            
            # 标记数据库中已存在的值
            field_label = self._get_field_labels([field_name]).get(field_name, field_name)
            for value in existing_values:
                for row_num in value_to_rows[value]:
                    if row_num not in errors:  # 避免覆盖内部重复错误
                        errors[row_num] = f"字段 '{field_label}' 值 '{value}' 已存在于数据库中"
        
        return errors

    async def _batch_query_existing(
            self, db: AsyncSession, field: str, values: List[str]
    ) -> Set[str]:
        """
        批量查询数据库中某字段已存在的值集合

        Args:
            db: 数据库会话
            field: 要查询的字段名
            values: 要检查的值列表

        Returns:
            数据库中已存在的值集合（字符串形式）
        """
        from sqlalchemy import text as sa_text

        if not values:
            return set()

        table = self.form_meta.main_table
        schema = await self._resolve_schema(db, self.form_meta.main_table_schema) or None
        database = self.form_meta.main_table_database or None
        table_name = self.sql_builder.build_table_name(table, schema, database)
        quoted_field = self.sql_builder.quote_identifier(field)

        existing: Set[str] = set()
        unique_values = list(dict.fromkeys(values))

        for i in range(0, len(unique_values), 500):
            batch_vals = unique_values[i:i + 500]
            placeholders = ", ".join(f":v{j}" for j in range(len(batch_vals)))
            sql = f"SELECT {quoted_field} FROM {table_name} WHERE {quoted_field} IN ({placeholders})"
            params = {f"v{j}": v for j, v in enumerate(batch_vals)}

            if await self._check_column_exists(db, table, schema, database, "is_deleted"):
                quoted_is_deleted = self.sql_builder.quote_identifier("is_deleted")
                sql += f" AND {quoted_is_deleted} = :is_deleted"
                params["is_deleted"] = False

            result = await db.execute(sa_text(sql), params)
            for row in result.fetchall():
                existing.add(str(row[0]))

        return existing

    def _validate_data_against_schema(self, row_num: int, data: Dict[str, Any]) -> Optional[str]:
        """
        根据数据库表结构验证数据
        
        Args:
            row_num: 行号
            data: 数据字典
        
        Returns:
            错误信息，如果验证通过则返回 None
        """
        form_config = self._form_config
        table_configs = form_config.get('tableConfigs', [])
        
        # 找到主表配置
        main_table_config = None
        for tc in table_configs:
            if tc.get('type') == 'main':
                main_table_config = tc
                break
        
        if not main_table_config:
            return None
        
        fields = main_table_config.get('fields', [])
        field_map = {f.get('name'): f for f in fields if f.get('name')}
        
        # 验证每个字段
        for field_name, value in data.items():
            # 跳过系统字段
            if field_name in {'id', 'sys_create_datetime', 'sys_update_datetime', 
                             'sys_creator_id', 'sys_modifier_id', 'sys_dept_id', 
                             'is_deleted', 'sort'}:
                continue
            
            field_config = field_map.get(field_name)
            if not field_config:
                continue
            
            # 1. NOT NULL 约束检查
            nullable = field_config.get('nullable', True)
            if not nullable and (value is None or (isinstance(value, str) and value.strip() == '')):
                return f"字段 '{field_name}' 不能为空"
            
            # 跳过 NULL 值的其他检查
            if value is None:
                continue
            
            # 2. 字符串长度检查
            field_type = field_config.get('type', '').lower()
            max_length = field_config.get('maxLength')
            
            if max_length and isinstance(value, str):
                if 'varchar' in field_type or 'char' in field_type or 'text' in field_type:
                    if len(value) > max_length:
                        return f"字段 '{field_name}' 长度超限（最大 {max_length}，实际 {len(value)}）"
            
            # 3. 数值类型检查
            if 'int' in field_type or 'integer' in field_type:
                if not isinstance(value, (int, float)):
                    try:
                        int(value)
                    except (ValueError, TypeError):
                        return f"字段 '{field_name}' 必须是整数"
            
            elif 'decimal' in field_type or 'numeric' in field_type or 'float' in field_type or 'double' in field_type:
                if not isinstance(value, (int, float)):
                    try:
                        float(value)
                    except (ValueError, TypeError):
                        return f"字段 '{field_name}' 必须是数值"
                
                # 精度检查
                precision = field_config.get('precision')
                scale = field_config.get('scale')
                if precision and isinstance(value, (int, float)):
                    value_str = str(value)
                    if '.' in value_str:
                        int_part, dec_part = value_str.split('.')
                        if len(int_part) + len(dec_part) > precision:
                            return f"字段 '{field_name}' 精度超限（最大 {precision}）"
                        if scale and len(dec_part) > scale:
                            return f"字段 '{field_name}' 小数位数超限（最大 {scale}）"
            
            # 4. 日期时间类型检查
            elif 'date' in field_type or 'time' in field_type:
                if not isinstance(value, (date, datetime)):
                    return f"字段 '{field_name}' 必须是日期时间类型"
            
            # 5. 布尔类型检查
            elif 'bool' in field_type:
                if not isinstance(value, bool):
                    return f"字段 '{field_name}' 必须是布尔类型"
        
        return None

    async def _batch_insert_raw(
            self, db: AsyncSession, table: str, schema: Optional[str],
            database: Optional[str], data_list: List[Dict[str, Any]]
    ) -> None:
        """
        真正的批量 INSERT：一条 SQL 插入多行数据
        
        Args:
            db: 数据库会话
            table: 表名
            schema: Schema 名
            database: 数据库名
            data_list: 数据列表（已标准化）
        """
        if not data_list:
            return
        
        columns = list(data_list[0].keys())
        cols = ", ".join(self.sql_builder.quote_identifier(c) for c in columns)
        full_table = self.sql_builder.build_table_name(table, schema, database)
        
        # 构建多行 VALUES，使用命名参数
        row_placeholders = []
        params = {}
        for row_idx, data in enumerate(data_list):
            placeholders = []
            for col in columns:
                param_name = f"p{row_idx}_{col}"
                placeholders.append(f":{param_name}")
                params[param_name] = data.get(col)
            row_placeholders.append(f"({', '.join(placeholders)})")
        
        sql = f"INSERT INTO {full_table} ({cols}) VALUES {', '.join(row_placeholders)}"
        await self._execute_command(db, sql, params)
    
    def _normalize_data_for_insert(self, data: Dict[str, Any]) -> Dict[str, Any]:
        """
        标准化数据类型，确保所有值都是数据库可接受的类型
        
        Args:
            data: 原始数据字典
        
        Returns:
            标准化后的数据字典
        """
        normalized = {}
        for key, value in data.items():
            if value is None:
                normalized[key] = None
            elif isinstance(value, datetime):
                # datetime 对象保持不变，数据库驱动会正确处理
                normalized[key] = value
            elif isinstance(value, date):
                # date 对象保持不变
                normalized[key] = value
            elif isinstance(value, (int, float, bool)):
                # 基本类型保持不变
                normalized[key] = value
            elif isinstance(value, (list, dict)):
                # 列表和字典转换为 JSON 字符串
                normalized[key] = json.dumps(value, ensure_ascii=False)
            elif isinstance(value, str):
                # 字符串保持不变
                normalized[key] = value
            else:
                # 其他类型转换为字符串
                normalized[key] = str(value)
        return normalized

    # ============ 字段权限过滤 ============

    async def apply_field_permissions(
            self,
            data: Any,
            db: AsyncSession,
            role_ids: List[str] = None
    ) -> Any:
        """
        应用字段权限过滤
        
        Args:
            data: 数据（单个对象或列表）
            db: 数据库会话
            role_ids: 角色ID列表，如果不传则从上下文获取
        
        Returns:
            过滤后的数据
        """
        from utils.context import get_current_user_info_from_context
        from core.resource_scope.field_permission.service import ResourceFieldPermissionService
        
        # 获取角色ID
        if not role_ids:
            user_info = get_current_user_info_from_context()
            logger.debug(f"[字段权限] 从上下文获取用户信息: {user_info}")
            if not user_info or not user_info.get('role_ids'):
                logger.warning("[字段权限] 未获取到用户角色信息，跳过字段权限过滤")
                return data
            role_ids = user_info['role_ids']
        
        # 获取资源类型
        resource_type = f"form:{self.form_meta.code}"
        logger.debug(f"[字段权限] 资源类型: {resource_type}, 角色IDs: {role_ids}")
        
        # 获取字段权限配置
        configs = await ResourceFieldPermissionService.get_by_roles_and_resource(
            db, role_ids, resource_type
        )
        logger.debug(f"[字段权限] 获取到的配置数量: {len(configs) if configs else 0}")
        
        if not configs:
            logger.debug("[字段权限] 未找到字段权限配置，跳过过滤")
            return data
        
        # 合并权限
        merged_perms = await ResourceFieldPermissionService.merge_field_permissions(
            configs, "most_permissive"
        )
        logger.debug(f"[字段权限] 合并后的权限: {merged_perms}")
        
        if not merged_perms:
            return data
        
        # 处理单个对象或列表
        if isinstance(data, list):
            return [self._apply_field_permissions(item, merged_perms) for item in data]
        else:
            return self._apply_field_permissions(data, merged_perms)

    def _apply_field_permissions(self, item: Dict[str, Any], field_perms: Dict[str, Dict]) -> Dict[str, Any]:
        """
        应用字段权限过滤（隐藏、脱敏）
        
        Args:
            item: 数据项（字典）
            field_perms: 字段权限配置
        
        Returns:
            过滤后的字典
        """
        if not isinstance(item, dict):
            return item
        
        # 收集需要隐藏或脱敏的字段及其关联的 _name 字段
        hidden_fields = set()
        masked_fields = {}  # field_name -> mask_rule
        
        for field_name, perm in field_perms.items():
            permission_type = perm.get('permission_type') or perm.get('permission', 'write')
            if permission_type == 'hidden':
                hidden_fields.add(field_name)
                # 同时隐藏关联的 _name 字段
                # 支持多种命名模式：
                # - field -> field_name (如 居住地 -> 居住地_name, post_id -> post_id_name)
                # - field_id -> field_name (如 manger_id -> manger_name)
                hidden_fields.add(f"{field_name}_name")
                if field_name.endswith('_id'):
                    base_name = field_name[:-3]  # 去掉 _id
                    hidden_fields.add(f"{base_name}_name")
            elif permission_type == 'masked':
                masked_fields[field_name] = perm.get('mask_rule')
                # 同时脱敏关联的 _name 字段（使用默认脱敏规则）
                masked_fields[f"{field_name}_name"] = 'default'
                if field_name.endswith('_id'):
                    base_name = field_name[:-3]
                    masked_fields[f"{base_name}_name"] = 'name'
        
        filtered = {}
        for field_name, value in item.items():
            # 检查是否需要隐藏
            if field_name in hidden_fields:
                continue
            
            # 检查是否需要脱敏
            if field_name in masked_fields:
                filtered[field_name] = self._mask_value(value, masked_fields[field_name])
                continue
            
            # 检查原始字段权限配置
            perm = field_perms.get(field_name, {})
            permission_type = perm.get('permission_type') or perm.get('permission', 'write')
            
            if permission_type == 'hidden':
                continue
            elif permission_type == 'masked':
                filtered[field_name] = self._mask_value(value, perm.get('mask_rule'))
            else:
                filtered[field_name] = value
        
        return filtered

    def _mask_value(self, value: Any, mask_rule: Optional[str]) -> str:
        """
        脱敏处理
        
        Args:
            value: 原始值
            mask_rule: 脱敏规则
        
        Returns:
            脱敏后的值
        """
        if not value:
            return value
        
        value_str = str(value)
        
        if mask_rule == "phone":
            # 手机号脱敏：138****5678
            if len(value_str) == 11:
                return f"{value_str[:3]}****{value_str[-4:]}"
        elif mask_rule == "email":
            # 邮箱脱敏：abc***@example.com
            if "@" in value_str:
                local, domain = value_str.split("@", 1)
                if len(local) > 3:
                    return f"{local[:3]}***@{domain}"
                return f"{local[0]}***@{domain}"
        elif mask_rule == "id_card":
            # 身份证脱敏：110***********1234
            if len(value_str) >= 8:
                return f"{value_str[:3]}***********{value_str[-4:]}"
        elif mask_rule == "name":
            # 姓名脱敏：张*
            if len(value_str) > 1:
                return f"{value_str[0]}*"
            return "*"
        
        # 默认脱敏：显示前后各2个字符
        if len(value_str) > 4:
            return f"{value_str[:2]}***{value_str[-2:]}"
        return "***"
